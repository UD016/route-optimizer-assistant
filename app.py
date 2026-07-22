# app.py
# Streamlit Route Optimizer — Page 1 + Page 2 (Planning)
# ─────────────────────────────────────────────────────
# PERFORMANCE FIXES applied (vs original):
#   [CRITIQUE-1] SQLite : connexion unique via @st.cache_resource (_get_db)
#                → plus de 400+ open/close par run
#   [CRITIQUE-2] get_job_pool_for_tech : haversine vectorisé NumPy
#                → plus de boucle iterrows O(n) avec geocode à chaque fois
#   [CRITIQUE-3] solo_jobs : set booked_ids (O(1)) au lieu de .copy() en boucle
#   [ÉLEVÉ-1]   _fetch_excel_df_from_github : bytes cachés séparément
#                → 1 seul GET HTTP même si les 2 fonctions appellent en même temps
#   [ÉLEVÉ-2]   tech_ll_map / tech_sector_map : @st.cache_data sur les coords techs
#   [ÉLEVÉ-3]   repair_month_plan : mini-cache local travel évite appels redondants
#   [ÉLEVÉ-4]   build_address : vectorisé str.cat (plus de apply axis=1)
#   [MOYEN-1]   duo_jobs : trié une seule fois avant la boucle, pas à chaque tour
#   [MOYEN-2]   Geotab : ThreadPoolExecutor pour fetch positions en parallèle
#   [MOYEN-3]   integrity check : sets stockés dans session_state (pas recalculés)
#   [MOYEN-4]   mode AUTO : binary search sur k au lieu de linéaire
#   [FAIBLE-1]  cummins_header : find_logo_path() cachée
#   [FAIBLE-2]  _norm_base/_normalize_base_job_id : une seule fonction au niveau module
#   [FAIBLE-3]  _choose_onsite_no_crumbs : une seule fonction au niveau module

"""
- Added service assistant integration
- Added conversation memory feature 
    - Each conversation clear leads to a clean slate
"""

import os
import re
import math
import calendar
import sqlite3
import time
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from datetime import datetime, date, timedelta, timezone
from typing import List, Optional, Tuple, Dict, Any

import numpy as np
import streamlit as st
import googlemaps
import polyline
import folium
from streamlit_folium import st_folium

import uuid

import pandas as pd
import requests

from service_assistant import ask_service_assistant

from timesheet import show_timesheet

from zoneinfo import ZoneInfo
TZ_LOCAL = ZoneInfo("America/Montreal")

# OR-Tools — optimisation de routes (pip install ortools)
try:
    from ortools.constraint_solver import routing_enums_pb2, pywrapcp
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False

# ────────────────────────────────────────────────────────────────
# Optional myGeotab import
# ────────────────────────────────────────────────────────────────
GEOTAB_AVAILABLE = True
try:
    import mygeotab as myg
except Exception:
    GEOTAB_AVAILABLE = False

# ────────────────────────────────────────────────────────────────
# Page config (ONE TIME)
# ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Route Optimizer", layout="wide", initial_sidebar_state="expanded")

# ────────────────────────────────────────────────────────────────
# Navigation persisted
# ────────────────────────────────────────────────────────────────
url_emp = st.query_params.get("emp", "").strip().upper()
is_tech = bool(url_emp)

if "page" not in st.session_state:
    st.session_state.page = "⏱ Feuille de temps" if is_tech else "🏠 Route Optimizer"

st.sidebar.title("Menu")

if not is_tech:
    # Superviseur — navigation complète
    st.session_state.page = st.sidebar.radio(
        "Navigation",
        ["🏠 Route Optimizer", "📅 Planning (Page 2)", "⏱ Feuille de temps"],
        index=["🏠 Route Optimizer", "📅 Planning (Page 2)", "⏱ Feuille de temps"].index(st.session_state.page),
        key="page_radio",
    )
else:
    # Tech — feuille de temps seulement, aucune navigation visible
    st.session_state.page = "⏱ Feuille de temps"

page = st.session_state.page
# ────────────────────────────────────────────────────────────────
# [FAIBLE-1] Header — logo path caché pour éviter 6x os.path.exists() par rerun
# ────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def find_logo_path() -> Optional[str]:
    candidates = [
        "Cummins_Logo.png", "Cummins_Logo.jpg", "Cummins_Logo.svg",
        "assets/cummins_black.svg", "assets/cummins_black.png", "assets/cummins_black.jpg",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None

def cummins_header():
    col_logo, col_title = st.columns([1, 5], vertical_alignment="center")
    with col_logo:
        logo_path = find_logo_path()
        if logo_path:
            try:
                st.image(logo_path, width=300)
            except Exception:
                logo_path = None
        if not logo_path:
            st.markdown(
                """
                <div style="width:150px;height:150px;display:flex;align-items:center;justify-content:center;">
                  <svg viewBox="0 0 100 100" width="150" height="150"
                       xmlns="http://www.w3.org/2000/svg" role="img" aria-label="Cummins">
                    <rect x="0" y="0" width="100" height="100" fill="#000000"/>
                    <path d="M70,50a20,20 0 1,1 -20,-20" fill="#ffffff"/>
                  </svg>
                </div>
                """,
                unsafe_allow_html=True
            )
    with col_title:
        st.markdown(
            """
            <div style="margin-bottom:2px;">
              <h1 style="margin:0;color:white;font-size:54px;">Optimisation du trajet des techniciens</h1>
            </div>
            <div style="color:#9aa0a6;font-size:32px;">
              Domicile ➜ Entrepôt ➜ Clients (MAXIMUM 25 TRAJETS) — <b>Cummins Service Fleet</b>
            </div>
            """,
            unsafe_allow_html=True
        )

# ────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────
def secret(name: str, default: Optional[str] = None) -> Optional[str]:
    try:
        return st.secrets[name]
    except Exception:
        return os.getenv(name, default)

def normalize_ca_postal(text: str) -> str:
    if not text:
        return text
    t = str(text).strip().upper().replace(" ", "")
    if len(t) == 6 and t[:3].isalnum() and t[3:].isalnum():
        return f"{t[:3]} {t[3:]}, Canada"
    return text

def big_number_marker(n: str, color_hex: str = "#cc3333"):
    html = f"""
    <div style="
      background:{color_hex};
      color:white;
      border-radius:18px;
      width:36px;height:36px;
      display:flex;align-items:center;justify-content:center;
      font-weight:700;font-size:16px;border:2px solid #222;">
      {n}
    </div>
    """
    return folium.DivIcon(html=html)

def recency_color(ts: Optional[str]) -> Tuple[str, str]:
    if not ts:
        return "#9e9e9e", "> 30d"
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except Exception:
        return "#9e9e9e", "unknown"
    age = datetime.now(timezone.utc) - dt.astimezone(timezone.utc)
    if age <= timedelta(hours=2):
        return "#00c853", "≤ 2h"
    if age <= timedelta(hours=24):
        return "#2e7d32", "≤ 24h"
    if age <= timedelta(days=7):
        return "#fb8c00", "≤ 7d"
    return "#9e9e9e", "> 7d"

# ────────────────────────────────────────────────────────────────
# Google Maps key
# ────────────────────────────────────────────────────────────────
GOOGLE_KEY = secret("GOOGLE_MAPS_API_KEY")

if GOOGLE_KEY:
    gmaps_client = googlemaps.Client(key=GOOGLE_KEY)
else:
    gmaps_client = None

# ────────────────────────────────────────────────────────────────
# Geocoding helpers (CACHED — inchangé)
# ────────────────────────────────────────────────────────────────
@st.cache_data(ttl=60*60*24*30, show_spinner=False, max_entries=20000)
def _geocode_cached(q: str) -> Optional[Tuple[float, float, str]]:
    if not q:
        return None
    try:
        res = gmaps_client.geocode(q, components={"country": "CA"}, region="ca")
        if res:
            loc = res[0]["geometry"]["location"]
            addr = res[0].get("formatted_address") or q
            return float(loc["lat"]), float(loc["lng"]), addr
    except Exception:
        pass
    return None

def geocode_ll(text: str) -> Optional[Tuple[float, float, str]]:
    if not text:
        return None
    q = normalize_ca_postal(text)
    return _geocode_cached(q)

@st.cache_data(ttl=60*60*24*30, show_spinner=False, max_entries=20000)
def _reverse_geocode_cached(lat: float, lon: float) -> str:
    try:
        res = gmaps_client.reverse_geocode((lat, lon))
        if res:
            return res[0].get("formatted_address", f"{lat:.5f},{lon:.5f}")
    except Exception:
        pass
    return f"{lat:.5f},{lon:.5f}"

def reverse_geocode(lat: float, lon: float) -> str:
    lat = round(float(lat), 5)
    lon = round(float(lon), 5)
    if lat == -0.0: lat = 0.0
    if lon == -0.0: lon = 0.0
    return _reverse_geocode_cached(lat, lon)

# ────────────────────────────────────────────────────────────────
# [FAIBLE-2] normalize_base_job_id — une seule fonction au niveau module
# (remplace _norm_base ET _normalize_base_job_id dupliquées)
# ────────────────────────────────────────────────────────────────
def normalize_base_job_id(jid: str) -> str:
    """Normalise un job_id: enlève PART X/Y et convertit float→int (ex: '347745.0' → '347745')"""
    s = str(jid).strip()
    # Convertir float string en int string: "347745.0" → "347745"
    try:
        if s.endswith('.0') and s[:-2].isdigit():
            s = s[:-2]
        elif '.' in s:
            f = float(s.split('(')[0].strip())
            if f == int(f):
                s = str(int(f)) + (s[s.index('('):] if '(' in s else '')
    except Exception:
        pass
    return re.sub(r"\s*\(PART\s+\d+/\d+\)\s*$", "", s).strip()

# ────────────────────────────────────────────────────────────────
# [FAIBLE-3] _choose_onsite_no_crumbs — une seule fonction au niveau module
# (était dupliquée dans schedule_month_with_duo ET dans mode A)
# ────────────────────────────────────────────────────────────────
def choose_onsite_no_crumbs(remaining_min: int, max_onsite_today: int, min_chunk: int) -> int:
    remaining_min = int(remaining_min)
    max_onsite_today = int(max_onsite_today)
    min_chunk = int(min_chunk)
    if max_onsite_today <= 0:
        return 0
    onsite_today = min(remaining_min, max_onsite_today)
    if onsite_today >= remaining_min:
        return onsite_today
    rem_after = remaining_min - onsite_today
    if onsite_today < min_chunk:
        return 0
    if rem_after < min_chunk:
        if remaining_min <= max_onsite_today:
            return remaining_min
        onsite_today = remaining_min - min_chunk
        onsite_today = min(onsite_today, max_onsite_today)
        if onsite_today < min_chunk:
            return 0
        rem_after = remaining_min - onsite_today
        if rem_after > 0 and rem_after < min_chunk:
            return 0
        return onsite_today
    return onsite_today

# ────────────────────────────────────────────────────────────────
# Shared data: TECH_HOME / ENTREPOTS
# ────────────────────────────────────────────────────────────────
TECH_HOME = {
    "Alain Duguay": "1110 rue Proulx, Les Cèdres, QC J7T 1E6",
    "Alexandre Pelletier Guay": "163 21e ave, Sabrevois, J0J 2G0",
    "Ali Reza-Sabour": "226 rue Felx, Saint-Clet, QC J0P 1S0",
    "David Robitaille": "1271 route des lac, saint-marcelline de kildare, QC J0K 2Y0",
    "Patrick Robitaille": "3365 ave laurier est, Montréal, QC H1X 1V3",
    "Benoit Charrette-Gosselin": "34 rue de la Digue, Saint-Jérome, QC, J7Y 5J1",
    "Benoit Larame": "12 rue de Beaudry, Mercier, J6R 2N7",
    "Christian Dubrueil": "31 rue des Roitelets, Delson, J5B 1T6",
    "Donald Lagace (IN SHOP)": "Montée Saint-Régis, Sainte-Catherine, QC, Canada",
    "Elie Rajotte-Lemay": "3700 Mnt du 4e Rang, Les Maskoutains, J0H 1S0",
    "Francois Racine": "80 rue de Beaujeu, Coteau-du-lac, J0P 1B0",
    "Fredy Diaz": "312 rue de Valcourt, Blainville, J7B 1H3",
    "George Yamna": "Rue René-Lévesque, Saint-Eustache, J7R 7L4",
    "Kevin Duranceau": "943 rue des Marquises, Beloeil, J3G 6T9",
    "Louis Lauzon": "5005 rue Domville, Saint-Hubert, J3Y 1Y2",
    "Martin Bourbonnière": "1444 rue de l'Orchidée, L'Assomption QC J5W 6B3",
    "Maxime Roy": "1407 3e Rue, Saint-Blaise-sur-Richelieu, QC J0J 1W0",
    "Michael Sulte": "2020 chem. De Covery Hill, Hinchinbrooke, QC, J0S 1E0",
    "Patrick Bellefleur": "222 rue Charles-Gadiou, L'Assomption, J5W 0J4",
    "Pier-Luc Cote": "143 rue Ashby, Marieville, J3M 1P2",
    "Sebastien Pepin (IN SHOP)": "Saint-Valentin, QC, Canada",
    "Sergio Mendoza": "791 Rue des Marquises, Beloeil, QC J3G 6M6",
}

ENTREPOTS = {
    "Candiac": "315 Liberté, Candiac, QC J5R 6Z7",
    "Assomption": "119 rue de la Commissaires, Assomption, QC, Canada",
    "Boisbriand": "5025 rue Ambroise-Lafortune, Boisbriand, QC, Canada",
    "Mirabel": "1600 Montée Guenette, Mirabel, QC, Canada",
}

# ────────────────────────────────────────────────────────────────
# Helper map labels (inchangé)
# ────────────────────────────────────────────────────────────────
def add_labeled_marker(fmap: folium.Map, lat: float, lon: float, label: str, kind: str):
    if kind == "wh":
        icon = folium.Icon(color="red", icon="building", prefix="fa")
    else:
        icon = folium.Icon(color="blue", icon="user", prefix="fa")

    folium.Marker([lat, lon], icon=icon, popup=folium.Popup(label, max_width=320), tooltip=label).add_to(fmap)

    folium.Marker(
        [lat, lon],
        icon=folium.DivIcon(
            icon_size=(260, 22),
            icon_anchor=(0, -18),
            html=f"""
            <div style="display:inline-block;padding:2px 6px;
                font-size:12px;font-weight:700;color:#111;
                background:rgba(255,255,255,.95);
                border:1px solid #ddd;border-radius:6px;
                box-shadow:0 1px 2px rgba(0,0,0,.25);white-space:nowrap;">
                {label}
            </div>
            """
        ),
    ).add_to(fmap)

# ────────────────────────────────────────────────────────────────
# PAGE 1 (Route Optimizer) — logique inchangée
# ────────────────────────────────────────────────────────────────

def render_page_1():

    if gmaps_client is None:

        st.warning(
            "Google Maps is disabled for this test."
        )

        return

    cummins_header()

    st.markdown("### Travel options")
    c1, c2 = st.columns([1.2, 1.2])
    with c1:
        st.markdown("**Travel mode:** Driving")
        leave_now = st.checkbox("Leave now", value=True, key="leave_now")
        round_trip = st.checkbox("Return to home at the end (round trip)?", value=True, key="round_trip")
    with c2:
        traffic_model = st.selectbox("Traffic model", ["best_guess", "pessimistic", "optimistic"], index=0, key="traffic_model")
        planned_date = st.date_input("Planned departure date", value=date.today(), disabled=leave_now, key="planned_date")
        planned_time = st.time_input("Planned departure time", value=datetime.now().time(), disabled=leave_now, key="planned_time")

    st.markdown("<hr style='margin:30px 0; border:1px solid #444;'>", unsafe_allow_html=True)

    if leave_now:
        departure_dt = datetime.now(TZ_LOCAL)
    else:
        departure_dt = datetime.combine(planned_date, planned_time, tzinfo=TZ_LOCAL)

    TECHNICIANS = sorted(TECH_HOME.keys())

    EXCEL_URL = "https://cummins365.sharepoint.com/:x:/r/sites/GRP_CC40846-AdministrationFSPG/Shared%20Documents/Administration%20FSPG/Info%20des%20techs%20pour%20booking/CapaciteTechs_CandiacEtOttawa.xlsx?d=wa4a6497bebb642849d640c57e4db82de&csf=1&web=1&e=8ltLaR"
    GITHUB_RAW_URL = "https://raw.githubusercontent.com/AR76F/route-optimizer/main/CapaciteTechs_CandiacEtOttawa.xlsx"

    hcol, bcol = st.columns([3, 2], vertical_alignment="center")
    with hcol:
        st.markdown("### 🧰 Technician capacities")
    with bcol:
        st.link_button("📎 Informations supplémentaires sur les techniciens", EXCEL_URL)

    st.caption("Choisis le type de service. On affiche les techniciens qui ont ce training **complété**.")

    if st.button("🔄 Recharger les données des trainings (GitHub)", key="refresh_trainings"):
        _get_excel_bytes_cached.clear()
        get_training_options.clear()
        get_not_completed_by_col.clear()

    # [ÉLEVÉ-1] Bytes du fichier Excel cachés séparément — 1 seul GET HTTP
    @st.cache_data(ttl=300, show_spinner=False)
    def _get_excel_bytes_cached(url: str) -> bytes:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        return r.content

    def _fetch_excel_df(raw_url: str, sheet: str, header=None) -> pd.DataFrame:
        content = _get_excel_bytes_cached(raw_url)
        return pd.read_excel(BytesIO(content), sheet_name=sheet, header=header, engine="openpyxl")

    def _norm_name(s: str) -> str:
        return " ".join(str(s or "").strip().lower().split())

    def _excel_col_to_idx(col_letter: str) -> int:
        col_letter = col_letter.strip().upper()
        idx = 0
        for ch in col_letter:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    SHEET_NAME = "Trainings"
    NAMES_COL_LETTER = "C"
    HEADER_ROW = 2
    TRAINING_COL_RANGE = ("H", "X")
    DATA_ROW_START = 3
    DATA_ROW_END = 22

    @st.cache_data(ttl=300, show_spinner=False)
    def get_training_options() -> list[tuple[str, int]]:
        df = _fetch_excel_df(GITHUB_RAW_URL, sheet=SHEET_NAME, header=None)
        r = HEADER_ROW - 1
        c_start = _excel_col_to_idx(TRAINING_COL_RANGE[0])
        c_end = _excel_col_to_idx(TRAINING_COL_RANGE[1])
        options = []
        for c in range(c_start, c_end + 1):
            val = df.iat[r, c] if (r < len(df) and c < df.shape[1]) else None
            label = str(val).strip() if val is not None and str(val).strip().lower() not in ("", "nan") else ""
            if label:
                options.append((label, c))
        return options

    @st.cache_data(ttl=300, show_spinner=False)
    def get_not_completed_by_col(training_col_idx: int) -> set:
        df = _fetch_excel_df(GITHUB_RAW_URL, sheet=SHEET_NAME, header=None)
        name_col_idx = _excel_col_to_idx(NAMES_COL_LETTER)
        r_start = max(0, DATA_ROW_START - 1)
        r_end = min(len(df) - 1, DATA_ROW_END - 1)
        sub = df.iloc[r_start:r_end + 1, [name_col_idx, training_col_idx]].copy()
        sub.columns = ["name", "status"]
        sub["status_norm"] = sub["status"].astype(str).str.strip().str.lower()
        not_completed_mask = sub["status_norm"].isin({"not completed", "notcompleted", "incomplete"})
        not_completed = sub[not_completed_mask]["name"].dropna()
        return {_norm_name(n) for n in not_completed.tolist()}

    def eligible_for(training_col_idx: int):
        not_ok_norm = get_not_completed_by_col(training_col_idx)
        return [t for t in TECHNICIANS if _norm_name(t) not in not_ok_norm]

    _training_pairs = get_training_options()
    _training_labels = ["(choisir)"] + [p[0] for p in _training_pairs]
    label_to_col = {label: col for (label, col) in _training_pairs}

    sel_training = st.selectbox("Type de service requis", _training_labels, index=0, key="tech_caps_training")
    if sel_training and sel_training != "(choisir)":
        col_idx = label_to_col.get(sel_training)
        techs = eligible_for(col_idx) if col_idx is not None else []
        if techs:
            st.success(f"{len(techs)} technicien(s) disponible(s) pour **{sel_training}**")
            for t in techs:
                st.write(f"• {t}")
        else:
            st.warning("Aucun technicien avec ce training complété.")

    st.markdown("---")
    st.subheader("🎯 Point de départ")

    if "route_start" not in st.session_state:
        st.session_state.route_start = ""
    if "storage_text" not in st.session_state:
        st.session_state.storage_text = ""

    tabs = st.tabs(["🚚 Live Fleet (Geotab)", "🏠 Technician Home"])

    # TAB 1 — GEOTAB LIVE FLEET
    with tabs[0]:
        G_DB = secret("GEOTAB_DATABASE")
        G_USER = secret("GEOTAB_USERNAME")
        G_PWD = secret("GEOTAB_PASSWORD")
        G_SERVER = secret("GEOTAB_SERVER", "my.geotab.com")
        geotab_enabled_by_secrets = GEOTAB_AVAILABLE and all([G_DB, G_USER, G_PWD])

        if geotab_enabled_by_secrets:
            if "geo_refresh_key" not in st.session_state:
                st.session_state.geo_refresh_key = 0
            if st.button("🔄 Rafraîchir Geotab maintenant", key="geo_refresh_btn"):
                st.session_state.geo_refresh_key += 1

            @st.cache_resource(show_spinner=False)
            def _geotab_api_cached(user, pwd, db, server):
                api = myg.API(user, pwd, db, server)
                api.authenticate()
                return api

            @st.cache_data(ttl=900, show_spinner=False)
            def _geotab_devices_cached(user, pwd, db, server):
                api = _geotab_api_cached(user, pwd, db, server)
                devs = api.call("Get", typeName="Device", search={"isActive": True}) or []
                return [{"id": d["id"], "name": d.get("name") or d.get("serialNumber") or "unit"} for d in devs]

            # [MOYEN-2] Fetch positions Geotab en parallèle avec ThreadPoolExecutor
            @st.cache_data(ttl=75, show_spinner=False)
            def _geotab_positions_for(api_params, device_ids, refresh_key):
                user, pwd, db, server = api_params
                api = _geotab_api_cached(user, pwd, db, server)

                def fetch_one(did):
                    try:
                        dsi = api.call("Get", typeName="DeviceStatusInfo", search={"deviceSearch": {"id": did}})
                        lat = lon = when = None
                        driver_name = None
                        if dsi:
                            row = dsi[0]
                            lat, lon = row.get("latitude"), row.get("longitude")
                            when = row.get("dateTime") or row.get("lastCommunicated") or row.get("workDate")
                            if (lat is None or lon is None) and isinstance(row.get("location"), dict):
                                lat = row["location"].get("y")
                                lon = row["location"].get("x")
                            drv = row.get("driver")
                            if isinstance(drv, dict):
                                driver_name = drv.get("name")
                        if lat is not None and lon is not None:
                            return {"deviceId": did, "lat": float(lat), "lon": float(lon),
                                    "when": when, "driverName": driver_name}
                        return {"deviceId": did, "error": "no_position"}
                    except Exception:
                        return {"deviceId": did, "error": "error"}

                results = []
                with ThreadPoolExecutor(max_workers=6) as ex:
                    futures = {ex.submit(fetch_one, did): did for did in device_ids}
                    for f in as_completed(futures):
                        results.append(f.result())
                return results

            DEVICE_TO_DRIVER_RAW = {
                "01942": "ALI-REZA SABOUR", "24735": "PATRICK BELLEFLEUR", "23731": "ÉLIE RAJOTTE-LEMAY",
                "19004": "GEORGES YAMNA", "22736": "MARTIN BOURBONNIÈRE", "23738": "PIER-LUC CÔTÉ",
                "24724": "LOUIS LAUZON", "23744": "BENOÎT CHARETTE", "23727": "FREDY DIAZ",
                "23737": "ALAIN DUGUAY", "23730": "BENOÎT LARAMÉE", "24725": "CHRISTIAN DUBREUIL",
                "23746": "MICHAEL SULTE", "24728": "FRANÇOIS RACINE", "23743": "ALEX PELLETIER-GUAY",
                "23745": "KEVIN DURANCEAU", "23739": "MAXIME ROY",
            }

            import json
            try:
                j = secret("GEOTAB_DEVICE_TO_DRIVER_JSON")
                if j:
                    DEVICE_TO_DRIVER_RAW.update(json.loads(j))
            except Exception:
                pass

            def _norm(s: str) -> str:
                return " ".join(str(s or "").strip().upper().split())

            NAME2DRIVER, ID2DRIVER = {}, {}
            for k, v in DEVICE_TO_DRIVER_RAW.items():
                nk = _norm(k)
                if not nk:
                    continue
                if len(nk) > 12 or ("-" in nk and any(c.isalpha() for c in nk)):
                    ID2DRIVER[nk] = v
                else:
                    NAME2DRIVER[nk] = v

            def _driver_from_mapping(device_id: str, device_name: str) -> Optional[str]:
                n_id, n_name = _norm(device_id), _norm(device_name)
                return NAME2DRIVER.get(n_name) or ID2DRIVER.get(n_id) or ID2DRIVER.get(n_name) or NAME2DRIVER.get(n_id)

            def _label_for_device(device_id: str, device_name: str, driver_from_api: Optional[str]) -> str:
                driver = driver_from_api or _driver_from_mapping(device_id, device_name) or "(no driver)"
                dev_label = device_name or device_id
                return f"{driver} — {dev_label}"

            devs = _geotab_devices_cached(G_USER, G_PWD, G_DB, G_SERVER)
            if not devs:
                st.info("Aucun appareil actif trouvé.")
            else:
                options, label2id = [], {}
                for d in devs:
                    lbl = _label_for_device(d["id"], d["name"], None)
                    options.append(lbl)
                    label2id[lbl] = d["id"]

                picked_labels = st.multiselect(
                    "Sélectionner un ou plusieurs véhicules/techniciens à afficher :",
                    sorted(options),
                    default=[],
                    key="geo_pick_labels",
                )
                wanted_ids = [label2id[lbl] for lbl in picked_labels]

                if wanted_ids:
                    pts = _geotab_positions_for((G_USER, G_PWD, G_DB, G_SERVER), tuple(wanted_ids), st.session_state.geo_refresh_key)
                    id2name = {d["id"]: d["name"] for d in devs}
                    valid = [p for p in pts if "lat" in p and "lon" in p]
                    if valid:
                        avg_lat = sum(p["lat"] for p in valid) / len(valid)
                        avg_lon = sum(p["lon"] for p in valid) / len(valid)
                        fmap = folium.Map(location=[avg_lat, avg_lon], zoom_start=8, tiles="cartodbpositron")

                        choice_labels = []
                        for p in valid:
                            device_id = p["deviceId"]
                            device_name = id2name.get(device_id, device_id)
                            label = _label_for_device(device_id, device_name, p.get("driverName"))
                            choice_labels.append(label)

                            color, lab = recency_color(p.get("when"))
                            folium.CircleMarker(
                                [p["lat"], p["lon"]],
                                radius=8, color="#222", weight=2,
                                fill=True, fill_color=color, fill_opacity=0.9
                            ).add_to(fmap)

                            folium.Marker(
                                [p["lat"], p["lon"]],
                                popup=folium.Popup(
                                    f"<b>{label}</b><br>Recency: {lab}<br>{p['lat']:.5f}, {p['lon']:.5f}",
                                    max_width=320
                                ),
                                tooltip=label,
                                icon=folium.DivIcon(
                                    icon_size=(240, 22),
                                    icon_anchor=(0, -18),
                                    html=f"""
                                    <div style="display:inline-block;padding:2px 6px;
                                        font-size:12px;font-weight:700;color:#111;
                                        background:rgba(255,255,255,.95);
                                        border:1px solid #ddd;border-radius:6px;
                                        box-shadow:0 1px 2px rgba(0,0,0,.25);white-space:nowrap;">
                                        {label.split(' — ')[0]}
                                    </div>"""
                                )
                            ).add_to(fmap)

                        st_folium(fmap, height=800, width=1800)

                        start_choice = st.selectbox("Utiliser comme point de départ :", ["(aucun)"] + choice_labels, index=0, key="geo_start_choice")
                        if start_choice != "(aucun)":
                            chosen = valid[choice_labels.index(start_choice)]
                            picked_addr = reverse_geocode(chosen["lat"], chosen["lon"])
                            st.session_state.route_start = picked_addr
                            st.success(f"Départ défini depuis **{start_choice}** → {picked_addr}")
                    else:
                        st.warning("Aucune position exploitable pour les éléments sélectionnés (essayez de rafraîchir).")
                else:
                    st.info("Sélectionnez au moins un véhicule/technicien pour afficher la carte.")
        else:
            st.info("Geotab désactivé. Ajoutez `GEOTAB_DATABASE`, `GEOTAB_USERNAME`, `GEOTAB_PASSWORD` dans les Secrets.")

    # TAB 2 — TECH HOMES + ENTREPOTS
    with tabs[1]:
        st.markdown("### 🏠 Domiciles des techniciens et entrepôts")
        show_map = st.checkbox("Afficher la carte (techniciens + entrepôts)", value=False, key="techhome_show_map")

        def _extract_postal(addr: str) -> str:
            if not addr:
                return ""
            m = re.search(r"\b([A-Z]\d[A-Z])\s?(\d[A-Z]\d)\b", str(addr).upper())
            return (m.group(1) + m.group(2)) if m else ""

        tech_home_df = pd.DataFrame(
            [{"tech_name": name, "home_address": addr, "postal": _extract_postal(addr)}
             for name, addr in TECH_HOME.items()]
        )
        st.session_state["tech_home"] = tech_home_df

        if show_map:
            try:
                tech_points = []
                for name, addr in TECH_HOME.items():
                    g = geocode_ll(addr)
                    if g:
                        lat, lon, formatted = g
                        tech_points.append({"name": name, "address": formatted, "lat": lat, "lon": lon})

                ent_points = []
                for ent_name, addr in ENTREPOTS.items():
                    g = geocode_ll(addr)
                    if g:
                        lat, lon, formatted = g
                        ent_points.append({"name": ent_name, "address": formatted, "lat": lat, "lon": lon})

                points_all = tech_points + ent_points
                if points_all:
                    avg_lat = sum(p["lat"] for p in points_all) / len(points_all)
                    avg_lon = sum(p["lon"] for p in points_all) / len(points_all)
                    fmap = folium.Map(location=[avg_lat, avg_lon], zoom_start=8, tiles="cartodbpositron")
                    for p in ent_points:
                        add_labeled_marker(fmap, p["lat"], p["lon"], f"🏭 {p['name']}", kind="wh")
                    for p in tech_points:
                        add_labeled_marker(fmap, p["lat"], p["lon"], p["name"], kind="tech")
                    st_folium(fmap, height=800, width=1800)
                else:
                    st.warning("Aucun point géocodé à afficher.")
            except Exception as e:
                st.error(f"Erreur lors du chargement de la carte : {e}")

        st.markdown("#### Sélectionner les sources de départ / fin")
        c1b, c2b = st.columns(2)
        with c1b:
            tech_choice = st.selectbox(
                "Technicien → définir comme **départ**",
                ["(choisir)"] + sorted(TECH_HOME.keys()),
                key="tech_choice_start_tab2"
            )
            if tech_choice != "(choisir)":
                st.session_state.route_start = TECH_HOME[tech_choice]
                st.success(f"Départ défini sur **{tech_choice}** — {TECH_HOME[tech_choice]}")
        with c2b:
            ent_choice = st.selectbox(
                "Entrepôt → définir comme **stockage**",
                ["(choisir)"] + sorted(ENTREPOTS.keys()),
                key="entrepot_choice_storage_tab2"
            )
            if ent_choice != "(choisir)":
                st.session_state.storage_text = ENTREPOTS[ent_choice]
                st.success(f"Stockage défini sur **Entrepôt — {ent_choice}** — {ENTREPOTS[ent_choice]}")

    if st.session_state.get("route_start"):
        st.info(f"📍 **Point de départ sélectionné :** {st.session_state.route_start}")

    st.markdown("### Route stops")
    start_text = st.text_input("Technician home (START)", key="route_start",
                               placeholder="e.g., 123 Main St, City, Province")
    storage_text = st.text_input("Storage location (first stop)", key="storage_text",
                                 placeholder="e.g., 456 Depot Rd, City, Province")
    stops_text = st.text_area("Other stops (one ZIP/postal code or full address per line)",
                              height=140, placeholder="H0H0H0\nG2P1L4\n…", key="stops_text")
    other_stops_input = [s.strip() for s in stops_text.splitlines() if s.strip()]

    st.markdown("---")
    if st.button("🧭 Optimize Route", type="primary", key="optimize_btn"):
        try:
            start_text = st.session_state.get("route_start", "").strip()
            storage_query = normalize_ca_postal(storage_text.strip()) if storage_text else ""
            other_stops_queries = [normalize_ca_postal(s.strip()) for s in other_stops_input if s.strip()]

            failures = []
            start_g = geocode_ll(start_text)
            if not start_g:
                failures.append(f"START: `{start_text}`")

            storage_g = geocode_ll(storage_query) if storage_query else None
            if storage_query and not storage_g:
                failures.append(f"STORAGE: `{storage_text}`")

            wp_raw = []
            if storage_query:
                wp_raw.append(("Storage", storage_query))
            for i, q in enumerate(other_stops_queries, start=1):
                wp_raw.append((f"Stop {i}", q))

            wp_geocoded: List[Tuple[str, str, Tuple[float, float]]] = []
            for label, q in wp_raw:
                g = geocode_ll(q)
                if not g:
                    failures.append(f"{label}: `{q}`")
                else:
                    lat, lon, addr = g
                    wp_geocoded.append((label, addr, (lat, lon)))

            if failures:
                st.error("I couldn't geocode some locations:\n\n- " + "\n- ".join(failures) +
                         "\n\nTip: use full street addresses if a postal code fails.")
                st.stop()

            def to_ll_str(ll: Tuple[float, float]) -> str:
                return f"{ll[0]:.7f},{ll[1]:.7f}"

            start_ll = (start_g[0], start_g[1])
            start_addr = start_g[2]

            wp_addrs = [addr for (_lbl, addr, _ll) in wp_geocoded]
            wp_llstr = [to_ll_str(ll) for (_lbl, _addr, ll) in wp_geocoded]

            if len(wp_llstr) > 23:
                st.error("Too many stops. Google allows up to **25 total** (origin + destination + waypoints).")
                st.stop()

            if st.session_state.get("round_trip", True):
                destination_addr = start_addr
                destination_llstr = to_ll_str(start_ll)
                waypoints_for_api = wp_llstr[:]
            else:
                if wp_llstr:
                    destination_addr = wp_addrs[-1]
                    destination_llstr = wp_llstr[-1]
                    waypoints_for_api = wp_llstr[:-1]
                else:
                    if storage_g:
                        destination_addr = storage_g[2]
                        destination_llstr = to_ll_str((storage_g[0], storage_g[1]))
                    else:
                        destination_addr = start_addr
                        destination_llstr = to_ll_str(start_ll)
                    waypoints_for_api = []

            wp_arg = (["optimize:true"] + waypoints_for_api) if waypoints_for_api else None

            directions = gmaps_client.directions(
                origin=to_ll_str(start_ll),
                destination=destination_llstr,
                mode="driving",
                waypoints=wp_arg,
                departure_time=departure_dt,
                traffic_model=st.session_state.get("traffic_model", "best_guess"),
            )

            if not directions:
                st.error("No route returned by Google Directions (driving). Try replacing postal codes with full addresses.")
                st.json({"origin": to_ll_str(start_ll), "destination": destination_llstr, "waypoints": waypoints_for_api})
                st.stop()

            if waypoints_for_api:
                order = directions[0].get("waypoint_order", list(range(len(waypoints_for_api))))
                ordered_wp_addrs = [wp_addrs[i] for i in order]
                if not st.session_state.get("round_trip", True) and wp_addrs:
                    ordered_wp_addrs.append(destination_addr)
            else:
                ordered_wp_addrs = [] if st.session_state.get("round_trip", True) else [destination_addr]

            visit_texts = [start_addr] + ordered_wp_addrs + ([start_addr] if st.session_state.get("round_trip", True) else [destination_addr])

            legs = directions[0].get("legs", [])
            total_dist_m = sum(leg.get("distance", {}).get("value", 0) for leg in legs)
            total_sec = sum((leg.get("duration_in_traffic") or leg.get("duration") or {}).get("value", 0) for leg in legs)
            km = total_dist_m / 1000.0 if total_dist_m else 0.0
            mins = total_sec / 60.0 if total_sec else 0.0

            per_leg = []
            current_dt = departure_dt
            for i, leg in enumerate(legs, start=1):
                dur = leg.get("duration_in_traffic") or leg.get("duration") or {}
                dur_sec = int(dur.get("value", 0))
                leg_mins = round(dur_sec / 60.0)
                dist_m = int(leg.get("distance", {}).get("value", 0))
                dist_km = dist_m / 1000.0
                current_dt = current_dt + timedelta(seconds=dur_sec)
                arr_str = current_dt.strftime("%H:%M")
                stop_addr = visit_texts[i] if i < len(visit_texts) else ""
                per_leg.append({"idx": i, "to": stop_addr, "dist_km": dist_km, "mins": leg_mins, "arrive": arr_str})

            st.session_state.route_result = {
                "visit_texts": visit_texts,
                "km": km,
                "mins": mins,
                "start_ll": start_ll,
                "wp_geocoded": wp_geocoded,
                "round_trip": st.session_state.get("round_trip", True),
                "overview": directions[0].get("overview_polyline", {}).get("points"),
                "per_leg": per_leg,
            }

        except Exception as e:
            st.error(f"Unexpected error: {type(e).__name__}: {e}")
            st.exception(e)

    res = st.session_state.get("route_result")
    if res:
        visit_texts = res["visit_texts"]
        km = res["km"]
        mins = res["mins"]
        start_ll = tuple(res["start_ll"])
        wp_geocoded = res["wp_geocoded"]
        round_trip_res = res["round_trip"]
        overview = res.get("overview")
        per_leg = res.get("per_leg", [])

        st.markdown("#### Optimized order (Driving)")
        for ix, addr in enumerate(visit_texts):
            if ix == 0:
                st.write(f"**START** — {addr}")
            elif ix == len(visit_texts) - 1:
                st.write(f"**END** — {addr}")
            else:
                st.write(f"**{ix}** — {addr}")

        if per_leg:
            st.markdown("#### Stop-by-stop timing")
            for leg in per_leg:
                st.write(f"**{leg['idx']}** → _{leg['to']}_  •  {leg['dist_km']:.1f} km  •  {leg['mins']} mins  •  **ETA {leg['arrive']}**")

        show_map2 = st.checkbox("Show map", value=False, key="route_show_map")
        if show_map2:
            try:
                fmap = folium.Map(location=[start_ll[0], start_ll[1]], zoom_start=9, tiles="cartodbpositron")
                if overview:
                    try:
                        path = polyline.decode(overview)
                        folium.PolyLine(path, weight=7, color="#2196f3", opacity=0.9).add_to(fmap)
                    except Exception:
                        pass

                folium.Marker(
                    start_ll,
                    icon=folium.Icon(color="green", icon="play", prefix="fa"),
                    popup=folium.Popup(f"<b>START</b><br>{visit_texts[0]}", max_width=260)
                ).add_to(fmap)

                addr2ll = {addr: ll for (_lbl, addr, ll) in wp_geocoded}
                for i, addr in enumerate(visit_texts[1:-1], start=1):
                    ll = addr2ll.get(addr)
                    if ll:
                        folium.Marker(
                            ll,
                            popup=folium.Popup(f"<b>{i}</b>. {addr}", max_width=260),
                            icon=big_number_marker(str(i))
                        ).add_to(fmap)

                end_addr = visit_texts[-1]
                end_ll = addr2ll.get(end_addr)
                if not end_ll:
                    g = geocode_ll(end_addr)
                    if g:
                        end_ll = (g[0], g[1])

                if end_ll:
                    folium.Marker(
                        end_ll,
                        icon=folium.Icon(color="red", icon="flag-checkered", prefix="fa"),
                        popup=folium.Popup(f"<b>{'END (Home)' if round_trip_res else 'END'}</b><br>{end_addr}", max_width=260)
                    ).add_to(fmap)

                st_folium(fmap, height=800, width=1800)
            except Exception as e:
                st.warning(f"Map rendering skipped: {e}")

        st.success(f"**Total distance:** {km:.1f} km • **Total time:** {mins:.0f} mins (live traffic)")


# Service Assistant Addition

def render_service_assistant():
    st.markdown("### 🤖 Service Coordinator Assistant")
    st.caption("Ask dispatch, technician selection, booking, troubleshooting, and invoicing questions.")

    if "assistant_messages" not in st.session_state:
        st.session_state.assistant_messages = [
            {"role": "assistant", "content": "Bonjour — Posez une question. Hi — Ask a question."}
        ]

    if "assistant_session_id" not in st.session_state:
        st.session_state.assistant_session_id = str(uuid.uuid4())

    clear_col, _ = st.columns([1, 5])
    with clear_col:
        if st.button("🗑️ Clear conversation", key="assistant_clear"):
            st.session_state.assistant_messages = [
                {"role": "assistant", "content": "Bonjour — Posez une question. Hi — Ask a question."}
            ]
            
            # Start a brand-new conversation
            st.session_state.assistant_session_id = str(uuid.uuid4())
            st.rerun()

    for msg in st.session_state.assistant_messages:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    question = st.chat_input("Ask a service question...", key="assistant_chat_input")

    if question:
        st.session_state.assistant_messages.append(
            {"role": "user", "content": question}
        )

        with st.chat_message("assistant"):
            with st.spinner("Analyzing request..."):
                try:
                    answer = ask_service_assistant(
                        question = question,
                        session_id = st.session_state.assistant_session_id,
                    )
                except Exception as e:
                    answer = f"Sorry — I hit an error: {type(e).__name__}: {e}"
                st.markdown(answer)

        st.session_state.assistant_messages.append(
            {"role": "assistant", "content": answer}
        )
        st.rerun()

# ────────────────────────────────────────────────────────────────
# PAGE 2 (Planning)
# ────────────────────────────────────────────────────────────────
def render_page_2():
    st.title("📅 Planning (Page 2)")

    tech_df = st.session_state.get("tech_home")
    if tech_df is None or len(tech_df) == 0:
        def _extract_postal(addr: str) -> str:
            if not addr:
                return ""
            m = re.search(r"\b([A-Z]\d[A-Z])\s?(\d[A-Z]\d)\b", str(addr).upper())
            return (m.group(1) + m.group(2)) if m else ""
        tech_df = pd.DataFrame(
            [{"tech_name": name, "home_address": addr, "postal": _extract_postal(addr)}
             for name, addr in TECH_HOME.items()]
        )
        st.session_state["tech_home"] = tech_df

    expected_cols = {"tech_name", "home_address"}
    if not expected_cols.issubset(set(tech_df.columns)):
        st.error("tech_home doit contenir tech_name et home_address.")
        st.stop()

    st.subheader("📤 Jobs – Upload Excel")
    uploaded = st.file_uploader("Upload ton fichier Excel jobs", type=["xlsx"], key="jobs_uploader")
    if uploaded:
        st.session_state["jobs_file_bytes"] = uploaded.getvalue()
        st.session_state["jobs_file_name"] = uploaded.name

    if "jobs_file_bytes" not in st.session_state:
        st.info("Upload un fichier Excel pour continuer (il sera conservé même si tu changes de page).")
        st.stop()

    data = BytesIO(st.session_state["jobs_file_bytes"])
    try:
        jobs_raw = pd.read_excel(data, sheet_name="Export", engine="openpyxl")
    except Exception:
        data.seek(0)
        jobs_raw = pd.read_excel(data, sheet_name=0, engine="openpyxl")

    st.caption(f"Jobs détectés: {len(jobs_raw)}")
    st.dataframe(jobs_raw.head(20), use_container_width=True)

    def pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        cols = {c.lower().strip(): c for c in df.columns}
        for cand in candidates:
            k = cand.lower().strip()
            if k in cols:
                return cols[k]
        return None

    COL_ORDER = pick_col(jobs_raw, ["ORDER #", "ORDER#", "Order", "Job ID", "WO", "Work Order"])
    COL_CUST  = pick_col(jobs_raw, ["CUST. #", "CUST #", "CUST#", "CUSTOMER #", "CUSTOMER#", "Customer #"])
    COL_ADDR1 = pick_col(jobs_raw, ["ADDRESS 1", "ADDRESS1", "Address 1"])
    COL_ADDR2 = pick_col(jobs_raw, ["ADDRESS 2", "ADDRESS2", "Address 2"])
    COL_ADDR3 = pick_col(jobs_raw, ["ADDRESS 3", "ADDRESS3", "Address 3"])
    COL_CITY  = pick_col(jobs_raw, ["SITE CITY", "CITY", "City"])
    COL_PROV  = pick_col(jobs_raw, ["SITE STATE", "STATE", "Province"])
    COL_POST  = pick_col(jobs_raw, ["SITE ZIP CODE", "ZIP", "POSTAL", "Postal Code"])
    COL_DESC  = pick_col(jobs_raw, ["PM SERVICE DESC.", "DESCRIPTION", "Service Desc", "Desc"])
    COL_UP    = pick_col(jobs_raw, ["UPCOMING SERVICES", "Upcoming Services"])
    COL_ONS   = pick_col(jobs_raw, ["ONSITE SRT HRS", "ONSITE HOURS", "ONSITE HRS"])
    COL_SRT   = pick_col(jobs_raw, ["SRT HRS", "SRT HOURS", "HRS"])
    COL_TECHN = pick_col(jobs_raw, ["# OF TECHS NEEDED", "TECHS NEEDED", "Nbr Techs"])
    COL_LAST_INSP = pick_col(jobs_raw, [
        "LAST INSPECTION", "Last Inspection", "LastInspection", "LAST_INSPECTION",
        "DERNIÈRE INSPECTION", "Derniere inspection", "Dernière inspection"
    ])
    COL_DIFF = pick_col(jobs_raw, ["DIFFERENCE", "Difference", "Diff", "ÉCART", "Ecart"])
    COL_UNIT = pick_col(jobs_raw, ["UNIT", "Unit", "UNITE", "Unité", "UNITE #", "UNIT #"])
    COL_SERIAL = pick_col(jobs_raw, ["SERIAL NUMBER", "Serial Number", "SERIAL", "S/N", "SN", "Serial"])
    COL_ALL_OPEN_WORK = pick_col(jobs_raw, ["ALL OPEN WORK", "All Open Work", "ALL OPEN WO"])

    if not COL_ORDER:
        st.error("Je ne trouve pas la colonne Job/Order (#). Assure-toi qu'elle existe dans ton export.")
        st.stop()

    def clean_id(x):
        try:
            if pd.isna(x):
                return ""
        except Exception:
            pass
        s = str(x).strip()
        if re.fullmatch(r"\d+\.0+", s):
            return s.split(".")[0]
        return s

    # [ÉLEVÉ-4] build_address vectorisé — plus de apply(axis=1)
    def build_address_vectorized(df: pd.DataFrame) -> pd.Series:
        addr_cols = [c for c in [COL_ADDR1, COL_ADDR2, COL_ADDR3, COL_CITY, COL_PROV, COL_POST] if c]
        if not addr_cols:
            return pd.Series([""] * len(df), index=df.index)
        parts = df[addr_cols].fillna("").astype(str).apply(lambda col: col.str.strip())
        def join_row(row):
            return ", ".join(v for v in row if v)
        return parts.apply(join_row, axis=1)

    def extract_postal(s: str) -> str:
        if not s:
            return ""
        m = re.search(r"\b([A-Z]\d[A-Z])\s?(\d[A-Z]\d)\b", str(s).upper())
        return (m.group(1) + m.group(2)) if m else ""

    def _clean_text(x):
        try:
            if pd.isna(x):
                return ""
        except Exception:
            pass
        s = str(x).strip()
        return "" if s.lower() in ("nan", "none") else s

    jobs = pd.DataFrame()
    jobs["job_id"] = jobs_raw[COL_ORDER].apply(clean_id)
    jobs["cust"] = jobs_raw[COL_CUST].apply(clean_id) if COL_CUST else ""
    jobs["address"] = build_address_vectorized(jobs_raw)

    desc = jobs_raw[COL_DESC].fillna("").astype(str) if COL_DESC else ""
    up   = jobs_raw[COL_UP].fillna("").astype(str) if COL_UP else ""
    jobs["description"] = (desc + " | " + up).str.strip(" |")

    ons = pd.to_numeric(jobs_raw[COL_ONS], errors="coerce") if COL_ONS else None
    srt = pd.to_numeric(jobs_raw[COL_SRT], errors="coerce") if COL_SRT else None
    if ons is not None:
        hours = ons
    elif srt is not None:
        hours = srt
    else:
        st.error("Je ne trouve pas ONSITE SRT HRS ni SRT HRS pour calculer la durée.")
        st.stop()

    jobs["job_minutes"] = (hours.fillna(0) * 60).round().astype(int)

    techs_needed = pd.to_numeric(jobs_raw[COL_TECHN], errors="coerce") if COL_TECHN else None
    jobs["techs_needed"] = techs_needed.fillna(1).astype(int) if techs_needed is not None else 1
    jobs["postal"] = jobs_raw[COL_POST].fillna("").astype(str).apply(extract_postal) if COL_POST else ""
    jobs["last_inspection"] = jobs_raw[COL_LAST_INSP].apply(_clean_text) if COL_LAST_INSP else ""
    jobs["difference"] = jobs_raw[COL_DIFF].apply(_clean_text) if COL_DIFF else ""
    jobs["unit"] = jobs_raw[COL_UNIT].apply(_clean_text) if COL_UNIT else ""
    jobs["serial_number"] = jobs_raw[COL_SERIAL].apply(_clean_text) if COL_SERIAL else ""
    jobs["all_open_work"] = jobs_raw[COL_ALL_OPEN_WORK].apply(_clean_text) if COL_ALL_OPEN_WORK else ""

    jobs = jobs[(jobs["address"].astype(str).str.len() > 8) & (jobs["job_minutes"] > 0)].copy()

    dedup_cols = ["job_id", "address", "job_minutes", "techs_needed", "description"]
    dedup_cols = [c for c in dedup_cols if c in jobs.columns]
    jobs = jobs.drop_duplicates(subset=dedup_cols).reset_index(drop=True)
    jobs = jobs.sort_values(["techs_needed", "job_id"], kind="mergesort").reset_index(drop=True)

    initial_bookable_job_ids = set(jobs["job_id"].astype(str).tolist())

    # ────────────────────────────────────────────────────────────────
    # Shared helpers
    # ────────────────────────────────────────────────────────────────
    MONTHS_FR = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    def month_selector(prefix: str) -> Tuple[date, date]:
        today = date.today()
        c1, c2 = st.columns(2)
        with c1:
            year = st.selectbox("Année", list(range(today.year - 1, today.year + 3)), index=1, key=f"{prefix}_year")
        with c2:
            month_name = st.selectbox("Mois", MONTHS_FR, index=today.month - 1, key=f"{prefix}_month")
        month_num = MONTHS_FR.index(month_name) + 1
        month_start = date(year, month_num, 1)
        last_day = calendar.monthrange(year, month_num)[1]
        month_end = date(year, month_num, last_day)
        st.caption(f"📅 Période planifiée : {month_start} → {month_end} (lundi→vendredi)")
        return month_start, month_end

    def business_days(start: date, end: date) -> List[date]:
        out = []
        d = start
        while d <= end:
            if d.weekday() < 5:
                out.append(d)
            d += timedelta(days=1)
        return out

    DAY_START_MIN = 480  # 08:00

    def mm_to_hhmm(m: int) -> str:
        total = int(m) + DAY_START_MIN
        h = total // 60
        mm = total % 60
        return f"{h:02d}:{mm:02d}"

    # ────────────────────────────────────────────────────────────────
    # [CRITIQUE-1] SQLite — connexion unique via @st.cache_resource
    # ────────────────────────────────────────────────────────────────
    from pathlib import Path
    DB_DIR = Path(".cache")
    DB_DIR.mkdir(exist_ok=True)
    DB_PATH = str(DB_DIR / "travel_cache.sqlite")

    @st.cache_resource
    def _get_db() -> sqlite3.Connection:
        """
        Connexion SQLite partagée pour toute la durée de vie de l'app.
        Créée une seule fois — jamais réouverte ni refermée.
        check_same_thread=False requis car Streamlit est multi-thread.
        """
        conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS travel (
                k TEXT PRIMARY KEY,
                minutes INTEGER,
                ts INTEGER
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_travel_ts ON travel(ts)")
        conn.commit()
        return conn

    st.sidebar.subheader("🧾 Coûts / Cache")
    use_traffic = st.sidebar.checkbox("Utiliser trafic (duration_in_traffic)", value=True, key="p2_use_traffic")
    cache_days = st.sidebar.number_input("Conserver cache (jours)", min_value=1, max_value=365, value=30, step=1, key="p2_cache_days")

    st.sidebar.subheader("🧭 Préfiltrage géographique")
    solo_pool = st.sidebar.slider("SOLO: nb de jobs candidats (pool)", 10, 300, 150, 10, key="p2_solo_pool")
    duo_pool = st.sidebar.slider("DUO: nb de jobs candidats (pool)", 10, 300, 80, 10, key="p2_duo_pool")
    techs_near_job = st.sidebar.slider("DUO: nb de techs proches à tester", 2, 6, 4, 1, key="p2_duo_near_techs")
    include_nearby_fsa = st.sidebar.checkbox("Inclure FSA voisins si peu de candidats", value=True, key="p2_include_nearby_fsa")

    st.sidebar.subheader("\U0001f5d1\ufe0f Cache zones géographiques")
    st.sidebar.caption("À utiliser si vous changez la logique des zones ou après une mise à jour.")
    if st.sidebar.button("\U0001f504 Recalculer zones géo (techs + jobs)", key="p2_reset_geo_cache"):
        st.session_state.pop("p2_ll_cache", None)
        # compute_tech_maps est défini plus bas - vider via session_state flag
        st.session_state["p2_reset_tech_maps"] = True
        st.success("\u2705 Cache zones vidé — secteurs recalculés au prochain run.")
        st.rerun()

    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s or "").strip().lower())

    def _key(origin: str, dest: str, traffic_flag: bool) -> str:
        raw = f"{_norm(origin)}|{_norm(dest)}|driving|traffic={int(bool(traffic_flag))}"
        return hashlib.md5(raw.encode("utf-8")).hexdigest()

    if "p2_api_calls" not in st.session_state:
        st.session_state["p2_api_calls"] = 0
    if "p2_cache_hits" not in st.session_state:
        st.session_state["p2_cache_hits"] = 0

    def travel_min_estimate(origin_addr: str, dest_addr: str,
                            origin_lat=None, origin_lon=None,
                            dest_lat=None, dest_lon=None) -> int:
        """
        Estimation rapide du temps de trajet via haversine × 1.5 (facteur urbain).
        ZÉRO appel API. Utilisé pendant l'évaluation des candidats dans le greedy.
        L'API réelle est appelée seulement au moment du booking final.
        Retourne 9999 si coordonnées inconnues.
        """
        try:
            # Essayer d'abord depuis le cache SQLite sans appel API
            k = _key(origin_addr, dest_addr, use_traffic)
            now = int(time.time())
            min_ts = now - int(cache_days) * 86400
            conn = _get_db()
            row = conn.cursor().execute(
                "SELECT minutes FROM travel WHERE k=? AND ts>=?", (k, min_ts)
            ).fetchone()
            if row:
                return int(row[0])
        except Exception:
            pass

        # Pas en cache → haversine
        try:
            if origin_lat is None or origin_lon is None:
                origin_lat, origin_lon = get_ll_for_address(origin_addr)
            if dest_lat is None or dest_lon is None:
                dest_lat, dest_lon = get_ll_for_address(dest_addr)
            if origin_lat and dest_lat:
                km = haversine_km(float(origin_lat), float(origin_lon),
                                  float(dest_lat), float(dest_lon))
                # Facteur 1.5 pour routes urbaines, minimum 5 min
                return max(5, int(km * 1.5))
        except Exception:
            pass
        return 60  # fallback raisonnable (1h) plutôt que 9999

    def travel_min_cached(origin: str, dest: str) -> int:
        """
        Retourne le temps de trajet en minutes entre origin et dest.
        Lit depuis SQLite si disponible (cache chaud).
        Sinon fait UN appel Distance Matrix 1x1 (fallback individuel).
        Pour le précalcul en masse, utiliser prefetch_travel_matrix().
        """
        if not origin or not dest:
            return 9999

        k = _key(origin, dest, use_traffic)
        now = int(time.time())
        min_ts = now - int(cache_days) * 86400

        conn = _get_db()
        cur = conn.cursor()
        cur.execute("SELECT minutes, ts FROM travel WHERE k=? AND ts>=?", (k, min_ts))
        row = cur.fetchone()
        if row:
            st.session_state["p2_cache_hits"] += 1
            return int(row[0])

        # Fallback 1x1 si pas en cache
        try:
            r = gmaps_client.distance_matrix([origin], [dest], mode="driving")
            el = r["rows"][0]["elements"][0]
            if el.get("status") != "OK":
                return 9999
            if use_traffic:
                dur = el.get("duration_in_traffic") or el.get("duration") or {}
            else:
                dur = el.get("duration") or el.get("duration_in_traffic") or {}
            minutes = int(round(int(dur.get("value", 0)) / 60))
        except Exception:
            return 9999

        conn.execute("INSERT OR REPLACE INTO travel(k, minutes, ts) VALUES(?,?,?)", (k, minutes, now))
        conn.commit()
        st.session_state["p2_api_calls"] += 1
        return minutes

    def prefetch_travel_matrix(origins: List[str], destinations: List[str],
                                progress_cb=None) -> int:
        """
        PRIORITÉ 1 — Distance Matrix BATCH.
        Précalcule et stocke dans SQLite toutes les paires (origin, dest)
        en appelant l'API Distance Matrix avec jusqu'à 25 origines × 25 destinations
        par requête, au lieu d'un appel séparé par paire.

        Exemple : 14 techs × 80 jobs = 1120 paires
          Avant : 1120 appels individuels
          Après : ceil(14/10) × ceil(80/25) = 2 × 4 = 8 appels batch
          → ~140x moins d'appels API

        Retourne le nombre de nouvelles paires calculées (pas déjà en cache).
        """
        if not origins or not destinations:
            return 0

        now = int(time.time())
        min_ts = now - int(cache_days) * 86400
        conn = _get_db()

        # Identifier les paires déjà en cache
        missing_origins: List[str] = []
        missing_dests_per_origin: Dict[str, List[str]] = {}

        for orig in origins:
            missing_dests = []
            for dest in destinations:
                if orig == dest:
                    continue
                k = _key(orig, dest, use_traffic)
                cur = conn.cursor()
                cur.execute("SELECT 1 FROM travel WHERE k=? AND ts>=?", (k, min_ts))
                if not cur.fetchone():
                    missing_dests.append(dest)
            if missing_dests:
                missing_origins.append(orig)
                missing_dests_per_origin[orig] = missing_dests

        if not missing_origins:
            return 0  # tout déjà en cache

        # Préparer les chunks : max 10 origines × 25 destinations par appel API
        # (Google limite à 100 éléments par requête = 10×10 ou 1×25, etc.)
        # On utilise 10 origines × 10 destinations = 100 éléments max
        CHUNK_ORIG = 10
        CHUNK_DEST = 10

        total_new = 0
        all_orig_chunks = [missing_origins[i:i+CHUNK_ORIG]
                           for i in range(0, len(missing_origins), CHUNK_ORIG)]

        total_calls = 0
        for orig_chunk in all_orig_chunks:
            # Collecter toutes les destinations manquantes pour ce chunk d'origines
            all_dests_for_chunk = list({d for o in orig_chunk
                                        for d in missing_dests_per_origin.get(o, [])})
            dest_chunks = [all_dests_for_chunk[i:i+CHUNK_DEST]
                           for i in range(0, len(all_dests_for_chunk), CHUNK_DEST)]

            for dest_chunk in dest_chunks:
                if not dest_chunk:
                    continue
                try:
                    r = gmaps_client.distance_matrix(
                        orig_chunk, dest_chunk, mode="driving"
                    )
                    total_calls += 1
                    rows_data = r.get("rows", [])

                    inserts = []
                    for oi, row_data in enumerate(rows_data):
                        if oi >= len(orig_chunk):
                            break
                        orig = orig_chunk[oi]
                        for di, el in enumerate(row_data.get("elements", [])):
                            if di >= len(dest_chunk):
                                break
                            dest = dest_chunk[di]
                            if orig == dest:
                                continue
                            # Vérifier que c'est une paire manquante pour cette origine
                            if dest not in missing_dests_per_origin.get(orig, []):
                                continue
                            if el.get("status") != "OK":
                                continue
                            if use_traffic:
                                dur = el.get("duration_in_traffic") or el.get("duration") or {}
                            else:
                                dur = el.get("duration") or el.get("duration_in_traffic") or {}
                            minutes = int(round(int(dur.get("value", 0)) / 60))
                            k = _key(orig, dest, use_traffic)
                            inserts.append((k, minutes, now))
                            total_new += 1

                    if inserts:
                        conn.executemany(
                            "INSERT OR REPLACE INTO travel(k, minutes, ts) VALUES(?,?,?)",
                            inserts
                        )
                        conn.commit()

                except Exception:
                    pass  # continuer même si un chunk échoue

                if progress_cb:
                    progress_cb(total_calls)

        st.session_state["p2_api_calls"] = st.session_state.get("p2_api_calls", 0) + total_calls
        return total_new

    # ────────────────────────────────────────────────────────────────
    # Haversine vectorisé (NumPy)
    # ────────────────────────────────────────────────────────────────
    def haversine_km(lat1, lon1, lat2, lon2) -> float:
        if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
            return 1e9
        R = 6371.0
        p1 = math.radians(float(lat1))
        p2 = math.radians(float(lat2))
        dp = math.radians(float(lat2) - float(lat1))
        dl = math.radians(float(lon2) - float(lon1))
        a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
        return 2 * R * math.asin(math.sqrt(a))

    def haversine_vectorized(tlat: float, tlon: float,
                              lats: np.ndarray, lons: np.ndarray) -> np.ndarray:
        """Calcule les distances haversine entre (tlat,tlon) et chaque point du vecteur."""
        R = 6371.0
        lats = np.where(np.isnan(lats.astype(float)), 1e9, lats.astype(float))
        lons = np.where(np.isnan(lons.astype(float)), 1e9, lons.astype(float))
        p1 = np.radians(float(tlat))
        p2 = np.radians(lats)
        dp = np.radians(lats - float(tlat))
        dl = np.radians(lons - float(tlon))
        a = np.sin(dp/2)**2 + np.cos(p1)*np.cos(p2)*np.sin(dl/2)**2
        a = np.clip(a, 0, 1)
        return 2 * R * np.arcsin(np.sqrt(a))

    if "p2_ll_cache" not in st.session_state:
        st.session_state["p2_ll_cache"] = {}
    ll_cache: Dict[str, Tuple[float, float]] = st.session_state["p2_ll_cache"]

    def _ensure_geocode_table():
        """Crée la table geocode dans SQLite si elle n'existe pas."""
        conn = _get_db()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS geocode (
                addr_key TEXT PRIMARY KEY,
                lat REAL,
                lon REAL,
                ts INTEGER
            )
        """)
        conn.commit()

    _ensure_geocode_table()

    # Pré-charger le geocode SQLite dans ll_cache au démarrage
    # (évite les appels API à chaque redémarrage Streamlit)
    if not ll_cache:
        try:
            conn = _get_db()
            rows = conn.execute("SELECT addr_key, lat, lon FROM geocode").fetchall()
            for addr_key, lat, lon in rows:
                ll_cache[addr_key] = (lat, lon)
        except Exception:
            pass

    def get_ll_for_address(addr: str) -> Tuple[Optional[float], Optional[float]]:
        if not addr:
            return None, None
        key = _norm(addr)
        # 1. Cache mémoire (session_state)
        if key in ll_cache:
            return ll_cache[key][0], ll_cache[key][1]
        # 2. Cache SQLite (persistant entre redémarrages)
        try:
            conn = _get_db()
            row = conn.execute("SELECT lat, lon FROM geocode WHERE addr_key=?", (key,)).fetchone()
            if row:
                ll_cache[key] = (row[0], row[1])
                return row[0], row[1]
        except Exception:
            pass
        # 3. Appel API Google (seulement si absent du cache)
        g = geocode_ll(addr)
        lat, lon = (float(g[0]), float(g[1])) if g else (None, None)
        ll_cache[key] = (lat, lon)
        # Persister dans SQLite
        try:
            conn = _get_db()
            conn.execute(
                "INSERT OR REPLACE INTO geocode (addr_key, lat, lon, ts) VALUES (?,?,?,?)",
                (key, lat, lon, int(__import__("time").time()))
            )
            conn.commit()
        except Exception:
            pass
        return lat, lon

    # ────────────────────────────────────────────────────────────────
    # ZONES GÉOGRAPHIQUES — 6 zones basées sur la géographie réelle
    # ────────────────────────────────────────────────────────────────
    #
    #  RIVE_NORD     : Nord du fleuve/Prairies, hors île MTL
    #                  Saint-Jérôme, Blainville, Laval, Repentigny,
    #                  L'Assomption, Terrebonne, Saint-Eustache, Lachute…
    #
    #  MTL_ILE       : Île de Montréal (zone neutre, compatible avec tout)
    #                  Montréal, Saint-Laurent, Dorval, Pointe-Claire,
    #                  Pierrefonds, Baie-d'Urfé, DDO, Verdun…
    #
    #  RIVE_SUD      : Sud du fleuve, corridor Montréal
    #                  Longueuil, Brossard, Boucherville, Chambly,
    #                  Beloeil, Varennes, Candiac, Saint-Hubert…
    #
    #  RIVE_SUD_OUEST: Sud-Ouest, accessible depuis Rive Nord via A-40
    #                  SANS traverser l'île de Montréal
    #                  Vaudreuil-Dorion, Saint-Clet, Hinchinbrooke,
    #                  Hudson, Rigaud, Beauharnois (ouest)…
    #
    #  ZONE_EST      : Centre/Est-du-Québec — Rive Sud SEULEMENT
    #                  Un tech Rive Nord doit traverser 2 fois l'île → INTERDIT
    #                  Drummondville, Sherbrooke, Windsor, Granby,
    #                  Waterloo, Bromont, Magog, Canton-de-Hatley…
    #
    #  ZONE_NORD     : Trois-Rivières et nord du fleuve hors corridor MTL
    #                  Accessible aux DEUX rives en longeant le fleuve
    #                  Trois-Rivières, Saint-Paulin, Shawinigan…
    #
    # ────────────────────────────────────────────────────────────────
    # MATRICE DE COMPATIBILITÉ (True = technicien PEUT faire ce job)
    #
    #                   RN   MTL   RS   RS_O  ZE   ZN
    # tech RIVE_NORD    ✅   ✅    ❌   ✅    ❌   ✅
    # tech MTL_ILE      ✅   ✅    ✅   ✅    ✅   ✅
    # tech RIVE_SUD     ❌   ✅    ✅   ✅    ✅   ✅
    # tech RS_OUEST     ❌   ✅    ✅   ✅    ✅   ✅
    # tech ZONE_EST     ❌   ✅    ✅   ❌    ✅   ❌
    # tech ZONE_NORD    ✅   ✅    ❌   ✅    ❌   ✅
    #
    # Règles clés :
    #   RIVE_NORD ↔ RIVE_SUD   : INTERDIT (traversée pont obligatoire)
    #   RIVE_NORD → ZONE_EST   : INTERDIT (2 ponts — traverse l'île MTL)
    #   RIVE_SUD  → RIVE_NORD  : INTERDIT (idem sens inverse)
    #   RIVE_SUD_OUEST         : accessible depuis Rive Nord via A-40
    #   ZONE_NORD (3R, SP)     : accessible des 2 rives le long du fleuve
    #   MTL_ILE                : zone neutre, compatible avec tout
    # ────────────────────────────────────────────────────────────────

    def classify_sector(lat: Optional[float], lon: Optional[float]) -> str:
        """
        Classifie une adresse dans l'une des 6 zones géographiques.
        Basé uniquement sur les coordonnées GPS — aucun appel API.

        Géographie réelle Montréal :
          Fleuve St-Laurent : lat ~45.40–45.55 selon longitude
          Île de Montréal   : lat 45.40–45.70, lon -74.03–-73.47
                              MAIS Laval est au NORD (lat > 45.62, nord de la Rivière-des-Prairies)
                              et Rive Sud est au SUD (lat < 45.50 hors île)
          Rivière-des-Prairies sépare Laval de l'île : lat ~45.62
        """
        if lat is None or lon is None:
            return "UNK"
        lat = float(lat)
        lon = float(lon)

        # ZONE_NORD — Trois-Rivières, Shawinigan, Saint-Paulin et nord
        if lat > 46.15:
            return "ZONE_NORD"

        # ZONE_EST — Drummondville, Sherbrooke, Granby, Bromont, Sorel, Waterloo…
        # Est de Montréal, lon > -73.15, inaccessible depuis Rive Nord (traversée obligatoire)
        if lon > -73.15 and 45.25 < lat < 46.10:
            return "ZONE_EST"

        # RIVE_SUD_OUEST — Vaudreuil, Saint-Clet, Rigaud, Hinchinbrooke…
        # Accessible depuis Rive Nord via A-40 SANS traverser l'île
        if lon < -74.00 and lat < 45.60:
            return "RIVE_SUD_OUEST"

        # Rive Sud à l'est de Longueuil — Varennes, Verchères, Contrecoeur
        # Le fleuve coule en diagonale : lat_fleuve ≈ 45.70 + (lon + 73.45)
        # Ces villes sont au sud du fleuve mais en latitudes parfois élevées
        if -73.45 < lon <= -73.15 and lat < (45.70 + (lon + 73.45)):
            return "RIVE_SUD"
        
        # RIVE_NORD stricte — lat > 45.70, hors île MTL
        # Blainville, Saint-Jérôme, Lachute, Terrebonne, Saint-Eustache, Repentigny…
        if lat > 45.70:
            return "RIVE_NORD"

        # LAVAL — île entre Rivière-des-Prairies et Montréal
        # lat 45.52–45.70, lon -73.97–-73.52
        # Traité comme RIVE_NORD : accessible depuis nord, PAS depuis Rive Sud directement
        if (45.52 <= lat <= 45.70) and (-73.97 <= lon <= -73.52):
            return "RIVE_NORD"

        # RIVE_NORD hors boîte — lat 45.62–45.70 à l'extérieur de l'île et de Laval
        # (ex: Lachute, Argenteuil à l'ouest)
        if lat >= 45.62 and not (-74.03 <= lon <= -73.38):
            return "RIVE_NORD"

        # MTL_ILE — Île de Montréal + rive proche (Brossard, Longueuil, Varennes…)
        # Étendu à l'est jusqu'à -73.38 pour capturer Varennes (rive sud est)
        if (45.42 <= lat <= 45.70) and (-74.03 <= lon <= -73.38):
            return "MTL_ILE"

        # RIVE_SUD — Sud du fleuve, corridor Montréal
        # Candiac, Saint-Hubert, Saint-Lambert, Chambly, Carignan, Beloeil…
        if lat < 45.55:
            return "RIVE_SUD"

        # Zone ambiguë restante → MTL_ILE par défaut
        return "MTL_ILE"

    # Matrice de compatibilité tech_zone → job_zone autorisés
    _SECTOR_COMPAT: Dict[str, set] = {
        # Règle fondamentale: personne au sud du fleuve ne traverse vers le nord
        # RIVE_NORD/ZONE_NORD (Tremblant, Saint-Jérôme, Lachute…) réservés
        # aux techs qui habitent déjà au nord du fleuve.
        "RIVE_NORD":      {"RIVE_NORD", "MTL_ILE", "RIVE_SUD_OUEST", "ZONE_NORD", "UNK"},
        "MTL_ILE":        {"MTL_ILE", "RIVE_SUD", "RIVE_SUD_OUEST", "ZONE_EST", "UNK"},
        "RIVE_SUD":       {"MTL_ILE", "RIVE_SUD", "RIVE_SUD_OUEST", "ZONE_EST", "UNK"},
        "RIVE_SUD_OUEST": {"MTL_ILE", "RIVE_SUD", "RIVE_SUD_OUEST", "ZONE_EST", "UNK"},
        "ZONE_EST":       {"MTL_ILE", "RIVE_SUD", "ZONE_EST", "UNK"},
        "ZONE_NORD":      {"RIVE_NORD", "MTL_ILE", "RIVE_SUD_OUEST", "ZONE_NORD", "UNK"},
        "UNK":            {"RIVE_NORD", "MTL_ILE", "RIVE_SUD", "RIVE_SUD_OUEST", "ZONE_EST", "ZONE_NORD", "UNK"},
    }

    def sector_compatible(tech_sector: str, job_sector: str) -> bool:
        """
        Retourne True si un technicien de tech_sector peut effectuer
        un job dans job_sector, selon la logique des ponts/traversées.
        """
        if tech_sector == "UNK" or job_sector == "UNK":
            return True
        return job_sector in _SECTOR_COMPAT.get(tech_sector, set())

    tech_names_all = sorted(tech_df["tech_name"].astype(str).tolist())
    home_map = {t: tech_df.loc[tech_df["tech_name"] == t, "home_address"].iloc[0] for t in tech_names_all}

    # [ÉLEVÉ-2] Coordonnées des techs cachées — recalculées uniquement si TECH_HOME change
    # Vider le cache si demandé via le bouton sidebar
    if st.session_state.pop("p2_reset_tech_maps", False):
        try:
            compute_tech_maps.clear()
        except Exception:
            pass

    @st.cache_data(show_spinner=False)
    def compute_tech_maps(home_map_items: tuple) -> Tuple[Dict, Dict]:
        """
        Calcule tech_ll_map et tech_sector_map une seule fois.
        home_map_items est un tuple de (name, address) pour permettre le cache.
        """
        t_ll: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
        t_sec: Dict[str, str] = {}
        for name, addr in home_map_items:
            g = geocode_ll(addr)
            if g:
                lat, lon, _ = g
                t_ll[name] = (float(lat), float(lon))
            else:
                t_ll[name] = (None, None)
            t_sec[name] = classify_sector(t_ll[name][0], t_ll[name][1])
        return t_ll, t_sec

    tech_ll_map, tech_sector_map = compute_tech_maps(tuple(sorted(home_map.items())))

    # ── Précalcul Distance Matrix Batch (sidebar) ─────────────────
    # Placé ICI car home_map et jobs sont maintenant définis
    st.sidebar.markdown("---")
    st.sidebar.subheader("⚡ Distance Matrix Batch")
    st.sidebar.caption(
        "Précalcule tous les trajets techs↔jobs en quelques appels API batch "
        "au lieu d'un appel par paire. À lancer UNE FOIS avant de planifier."
    )
    if st.sidebar.button("🔄 Précalculer matrice trajets", key="p2_prefetch_btn"):
        all_tech_addrs = list(home_map.values())
        all_job_addrs  = jobs["address"].dropna().unique().tolist()
        all_addrs      = list(dict.fromkeys(all_tech_addrs + all_job_addrs))

        prog_bar  = st.sidebar.progress(0)
        prog_text = st.sidebar.empty()
        CHUNK_SIZE = 100

        total_calls_est = max(1, (len(all_addrs) * len(all_addrs)) // CHUNK_SIZE)

        def _cb(calls_done):
            pct = min(99, int(calls_done / max(1, total_calls_est) * 100))
            prog_bar.progress(pct)
            prog_text.write(f"Appels batch effectués : {calls_done}")

        new_pairs = prefetch_travel_matrix(all_addrs, all_addrs, progress_cb=_cb)
        prog_bar.progress(100)
        prog_text.write(f"✅ {new_pairs} nouvelles paires ajoutées au cache.")
        st.sidebar.success(
            f"Matrice précalculée — {new_pairs} nouvelles paires. "
            f"Appels API batch : {st.session_state.get('p2_api_calls', 0)}"
        )

    st.sidebar.caption(
        "💡 OR-Tools actif" if ORTOOLS_AVAILABLE
        else "⚠️ OR-Tools non installé (`pip install ortools`). "
             "Greedy utilisé à la place."
    )

    if "job_lat" not in jobs.columns:
        jobs["job_lat"] = None
        jobs["job_lon"] = None
        jobs["job_sector"] = "UNK"

    def ensure_job_ll_master(master_df: pd.DataFrame, master_idx) -> Tuple[Optional[float], Optional[float], str]:
        r = master_df.loc[master_idx]
        lat = r.get("job_lat")
        lon = r.get("job_lon")
        sec = r.get("job_sector", "UNK") or "UNK"
        if pd.notna(lat) and pd.notna(lon):
            lat = float(lat); lon = float(lon)
            if sec == "UNK":
                sec = classify_sector(lat, lon)
                master_df.at[master_idx, "job_sector"] = sec
            return lat, lon, sec
        lat, lon = get_ll_for_address(r["address"])
        sec = classify_sector(lat, lon)
        master_df.at[master_idx, "job_lat"] = lat
        master_df.at[master_idx, "job_lon"] = lon
        master_df.at[master_idx, "job_sector"] = sec
        return lat, lon, sec

    # [CRITIQUE-2] get_job_pool_for_tech — haversine vectorisé NumPy
    def get_job_pool_for_tech(master_remaining: pd.DataFrame, tech_name: str, pool_size: int) -> pd.DataFrame:
        if master_remaining.empty:
            return master_remaining

        master_remaining = master_remaining.sort_values(["job_id"], kind="mergesort")
        tlat, tlon = tech_ll_map.get(tech_name, (None, None))
        tsec = tech_sector_map.get(tech_name, "UNK")

        if tlat is None or tlon is None:
            return master_remaining.head(pool_size) if len(master_remaining) > pool_size else master_remaining

        # S'assurer que les lat/lon des jobs sont disponibles dans le master df
        for idx in master_remaining.index:
            if idx in jobs.index:
                lat_val = jobs.at[idx, "job_lat"]
                if pd.isna(lat_val) or lat_val is None:
                    ensure_job_ll_master(jobs, idx)

        # Extraire les vecteurs lat/lon
        lats_raw = []
        lons_raw = []
        valid_idx = []
        for idx, row in master_remaining.iterrows():
            if idx in jobs.index:
                lat = jobs.at[idx, "job_lat"]
                lon = jobs.at[idx, "job_lon"]
                sec = jobs.at[idx, "job_sector"] if "job_sector" in jobs.columns else "UNK"
            else:
                lat, lon = get_ll_for_address(str(row.get("address", "")))
                sec = classify_sector(lat, lon)
            if not sector_compatible(tsec, sec or "UNK"):
                continue
            lats_raw.append(float(lat) if lat is not None else np.nan)
            lons_raw.append(float(lon) if lon is not None else np.nan)
            valid_idx.append(idx)

        if not valid_idx:
            return master_remaining.head(pool_size) if len(master_remaining) > pool_size else master_remaining

        lats_arr = np.array(lats_raw, dtype=float)
        lons_arr = np.array(lons_raw, dtype=float)
        dists = haversine_vectorized(tlat, tlon, lats_arr, lons_arr)

        top_n = min(int(pool_size), len(valid_idx))
        top_positions = np.argpartition(dists, top_n - 1)[:top_n] if top_n < len(dists) else np.arange(len(dists))
        chosen_idx = [valid_idx[i] for i in top_positions]
        return master_remaining.loc[chosen_idx].copy()

    def rank_techs_for_job(tech_names: List[str], job_row: pd.Series, top_n: int) -> List[str]:
        if top_n >= len(tech_names):
            return tech_names
        addr = job_row.get("address", "")
        # Utiliser ll_cache directement si disponible — évite appel geocode API
        _addr_key = re.sub(r"\s+", " ", str(addr or "").strip().lower())
        if _addr_key in ll_cache:
            jlat, jlon = ll_cache[_addr_key]
        else:
            jlat, jlon = get_ll_for_address(addr)
        jsec = classify_sector(jlat, jlon)
        scored = []
        for t in tech_names:
            tlat, tlon = tech_ll_map.get(t, (None, None))
            tsec = tech_sector_map.get(t, "UNK")
            if not sector_compatible(tsec, jsec):
                continue
            d = haversine_km(tlat, tlon, jlat, jlon)
            scored.append((t, d))
        if not scored:
            return tech_names[:max(2, int(top_n))]
        scored.sort(key=lambda x: x[1])
        return [t for (t, _d) in scored[:max(2, int(top_n))]]

    # ────────────────────────────────────────────────────────────────
    # Styling + Filters
    # ────────────────────────────────────────────────────────────────
    def style_duo(df: pd.DataFrame):
        if df is None or df.empty:
            return df

        def _row_style(row):
            ot_val = str(row.get("ot", "")).strip().lower()
            if ot_val:
                css = (
                    "background-color: #f8d7da;"
                    "color: #000000;"
                    "font-weight: 900;"
                    "border-left: 6px solid #b02a37;"
                )
                return [css] * len(row)
            try:
                n = int(row.get("techs_needed", 1))
            except Exception:
                n = 1
            if n >= 3:
                css = (
                    "background-color: #f8d7da;"
                    "color: #000000;"
                    "font-weight: 800;"
                    "border-left: 6px solid #b02a37;"
                )
                return [css] * len(row)
            if n == 2:
                css = (
                    "background-color: #ffd966;"
                    "color: #000000;"
                    "font-weight: 700;"
                    "border-left: 6px solid #ff9800;"
                )
                return [css] * len(row)
            return [""] * len(row)

        return df.style.apply(_row_style, axis=1)

    INSPECTION_KEYWORDS = ["inspection", "generator inspection", "génératrice inspection", "inspection génératrice"]

    def filter_by_service_type(df: pd.DataFrame, mode_label: str) -> pd.DataFrame:
        if mode_label == "Inclure full service (tous les jobs)":
            return df
        if "description" not in df.columns:
            return df
        s = df["description"].fillna("").astype(str).str.lower()
        mask = False
        for kw in INSPECTION_KEYWORDS:
            mask = mask | s.str.contains(kw, na=False)
        if mode_label == "Generator inspection seulement":
            return df[mask].copy()
        if mode_label == "Exclure generator inspection":
            return df[~mask].copy()
        return df

    def compute_total_parts(job_minutes_total: int, daily_onsite_cap: int) -> int:
        if daily_onsite_cap <= 0:
            return 1
        return int(math.ceil(float(job_minutes_total) / float(daily_onsite_cap)))


    # ════════════════════════════════════════════════════════════════
    # Scheduler principal
    # ════════════════════════════════════════════════════════════════
    def schedule_month_with_duo(
        jobs_in: pd.DataFrame,
        tech_names: List[str],
        month_days: List[date],
        day_hours: float,
        lunch_min: int,
        buffer_job: int,
        max_jobs_per_day: int,
        allow_duo: bool,
        progress=None,
        progress_text=None
    ) -> dict:

        available = int(round(day_hours * 60)) - int(lunch_min)
        if available <= 0:
            return {"success": False, "rows": [], "remaining": jobs_in, "reason": "Heures/jour - pause <= 0"}
        daily_onsite_cap = int(available)
        MIN_ONSITE_CHUNK_MIN = 180

        OT_ACTIVE_CAP = int(round(14 * 60)) - int(lunch_min)
        if OT_ACTIVE_CAP < available:
            OT_ACTIVE_CAP = available

        remaining_all = jobs_in.copy()
        remaining_all = remaining_all.sort_values(["techs_needed", "job_id"], kind="mergesort")

        duo_jobs = remaining_all[remaining_all["techs_needed"] == 2].copy() if allow_duo else remaining_all.iloc[0:0].copy()
        solo_jobs = remaining_all[remaining_all["techs_needed"] <= 1].copy()
        hard_jobs = remaining_all[remaining_all["techs_needed"] > 2].copy()

        # [MOYEN-1] Trier une seule fois avant les boucles
        duo_jobs = duo_jobs.sort_values(["job_id"], kind="mergesort")

        # Trier solo_jobs par contrainte géographique croissante :
        # Les jobs accessibles par peu de techs (zones spécifiques) passent EN PREMIER
        # pour éviter que le greedy remplisse les techs sur des jobs faciles
        # et laisse les jobs contraints pour la fin quand il n'y a plus de place.
        def _count_compatible_techs(row):
            try:
                addr = str(row.get("address", ""))
                lat = row.get("job_lat", None) if "job_lat" in row.index else None
                lon = row.get("job_lon", None) if "job_lon" in row.index else None
                if lat is None or pd.isna(lat):
                    lat, lon = get_ll_for_address(addr)
                sec = classify_sector(lat, lon)
                return sum(1 for t in tech_names if sector_compatible(_tech_sector.get(t,"UNK"), sec))
            except Exception:
                return len(tech_names)

        try:
            solo_jobs = solo_jobs.copy()
            solo_jobs["_n_techs_compat"] = solo_jobs.apply(_count_compatible_techs, axis=1)
            solo_jobs = solo_jobs.sort_values(["_n_techs_compat", "job_id"], kind="mergesort")
            solo_jobs = solo_jobs.drop(columns=["_n_techs_compat"])
        except Exception:
            solo_jobs = solo_jobs.sort_values(["job_id"], kind="mergesort")
        hard_jobs = hard_jobs.sort_values(["job_id"], kind="mergesort")

        planned_rows: List[Dict[str, Any]] = []
        planned_base_ids: set = set()

        def _extra_fields_from_job(job_like):
            getter = job_like.get if hasattr(job_like, "get") else (lambda _k, _d="": _d)
            return {
                "last_inspection": getter("last_inspection", ""),
                "difference": getter("difference", ""),
                "unit": getter("unit", ""),
                "serial_number": getter("serial_number", ""),
                "all_open_work": getter("all_open_work", ""),
            }

        _home_map = {t: home_map[t] for t in tech_names}
        _tech_sector = {t: tech_sector_map.get(t, "UNK") for t in tech_names}

        carryover_by_tech: Dict[str, Dict[str, Any]] = {}
        split_label_state: Dict[str, Dict[str, Any]] = {}

        def _register_and_relabel_split_row(base_job_id: str, new_row_idx: int, part_num: int):
            if base_job_id not in split_label_state:
                split_label_state[base_job_id] = {"total": int(part_num), "row_idxs": []}
            stt = split_label_state[base_job_id]
            if int(part_num) > int(stt["total"]):
                stt["total"] = int(part_num)
            stt["row_idxs"].append(int(new_row_idx))
            total = int(stt["total"])
            for i, ridx in enumerate(stt["row_idxs"], start=1):
                base = normalize_base_job_id(planned_rows[ridx]["job_id"])
                planned_rows[ridx]["job_id"] = f"{base} (PART {i}/{total})"

        def _book_split_part_for_tech(day, t, used, cur_loc, jobs_count, lock_tech, split_state):
            addr = split_state["address"]
            base_job_id = split_state["base_job_id"]
            cust = split_state.get("cust", "")
            desc = split_state.get("description", "")
            techs_needed_val = int(split_state.get("techs_needed", 1))

            tsec = _tech_sector.get(t, "UNK")
            jlat, jlon = get_ll_for_address(addr)
            jsec = classify_sector(jlat, jlon)
            if not sector_compatible(tsec, jsec):
                return

            remaining_min = int(split_state["remaining_job_min"])
            part_idx = int(split_state["part_idx_next"])

            tmin = travel_min_cached(cur_loc[t], addr)
            tback = travel_min_cached(addr, _home_map[t])

            max_onsite_today = int(available) - int(used[t]) - int(tmin) - int(buffer_job) - int(tback)
            if max_onsite_today <= 0:
                return

            onsite_today = choose_onsite_no_crumbs(remaining_min, max_onsite_today, MIN_ONSITE_CHUNK_MIN)
            if onsite_today <= 0:
                return

            jobs_count[t] += 1
            start_m = int(used[t]) + int(tmin)
            end_m = start_m + int(onsite_today) + int(buffer_job)

            planned_rows.append({
                "date": day.isoformat(),
                "technicien": t,
                "sequence": jobs_count[t],
                "job_id": f"{base_job_id} (PART {part_idx}/{max(1, int(split_state.get('total_parts', part_idx)))})",
                "cust": cust,
                "duo": "",
                "ot": "",
                "debut": mm_to_hhmm(int(start_m)),
                "fin": mm_to_hhmm(int(end_m)),
                "adresse": addr,
                "travel_min": int(tmin),
                "job_min": int(onsite_today),
                "buffer_min": int(buffer_job),
                "techs_needed": techs_needed_val,
                **_extra_fields_from_job(split_state),
                "description": desc,
            })
            planned_base_ids.add(normalize_base_job_id(base_job_id))

            row_idx = len(planned_rows) - 1
            _register_and_relabel_split_row(base_job_id, row_idx, part_idx)

            used[t] = int(end_m)
            cur_loc[t] = addr

            remaining_min = int(remaining_min) - int(onsite_today)
            split_state["remaining_job_min"] = remaining_min
            split_state["part_idx_next"] = part_idx + 1
            # Locker seulement si carryover encore actif ET pas assez de temps
            # pour un autre job après (évite de bloquer le tech pour la journée)
            time_left = int(available) - int(end_m)
            if remaining_min > 0 and time_left < int(MIN_ONSITE_CHUNK_MIN):
                lock_tech[t] = True
            else:
                lock_tech[t] = False

        total_steps = max(1, len(month_days))

        def _sort_techs_by_proximity(tech_list, remaining_jobs):
            """
            Trie les techs par distance haversine à leur job compatible le plus proche.
            Les techs avec un job très proche passent en premier → assignation naturelle
            par zone géographique (ex: David proche Tremblant prend les jobs Tremblant).
            """
            if remaining_jobs.empty:
                return tech_list
            scores = []
            for t in tech_list:
                tlat, tlon = tech_ll_map.get(t, (None, None))
                if tlat is None or tlon is None:
                    scores.append((t, float('inf')))
                    continue
                tsec = _tech_sector.get(t, "UNK")
                best_dist = float('inf')
                for idx, job in remaining_jobs.iterrows():
                    jlat = jobs.at[idx, "job_lat"] if idx in jobs.index and "job_lat" in jobs.columns else None
                    jlon = jobs.at[idx, "job_lon"] if idx in jobs.index and "job_lon" in jobs.columns else None
                    if jlat is None or pd.isna(jlat):
                        continue
                    jsec = jobs.at[idx, "job_sector"] if idx in jobs.index and "job_sector" in jobs.columns else "UNK"
                    if not sector_compatible(tsec, jsec or "UNK"):
                        continue
                    d = haversine_km(tlat, tlon, float(jlat), float(jlon))
                    if d < best_dist:
                        best_dist = d
                scores.append((t, best_dist))
            scores.sort(key=lambda x: x[1])
            return [t for t, _ in scores]

        for di, day in enumerate(month_days):
            _t_day_start = time.time()
            used = {t: 0 for t in tech_names}
            cur_loc = {t: _home_map[t] for t in tech_names}
            jobs_count = {t: 0 for t in tech_names}
            lock_tech = {t: False for t in tech_names}

            # Reconstruire solo_jobs depuis la source en excluant ce qui est déjà planifié
            # IMPORTANT: NE PAS reset_index — les index doivent correspondre à jobs.index
            solo_jobs = remaining_all[remaining_all["techs_needed"] <= 1].copy()
            solo_jobs = solo_jobs[
                ~solo_jobs["job_id"].apply(normalize_base_job_id).isin(planned_base_ids)
            ]
            try:
                solo_jobs["_n_techs_compat"] = solo_jobs.apply(_count_compatible_techs, axis=1)
                solo_jobs = solo_jobs.sort_values(["_n_techs_compat", "job_id"], kind="mergesort")
                solo_jobs = solo_jobs.drop(columns=["_n_techs_compat"])
            except Exception:
                solo_jobs = solo_jobs.sort_values(["job_id"], kind="mergesort")

            duo_jobs = remaining_all[remaining_all["techs_needed"] == 2].copy() if allow_duo else remaining_all.iloc[0:0].copy()
            duo_jobs = duo_jobs[
                ~duo_jobs["job_id"].apply(normalize_base_job_id).isin(planned_base_ids)
            ]

            for t in tech_names:
                if t in carryover_by_tech:
                    _book_split_part_for_tech(day, t, used, cur_loc, jobs_count, lock_tech, carryover_by_tech[t])
                    if int(carryover_by_tech[t]["remaining_job_min"]) <= 0:
                        del carryover_by_tech[t]

            # ---- 1) DUO first ----
            if allow_duo and (not duo_jobs.empty) and len(tech_names) >= 2:
                # [MOYEN-1] Pas de re-tri à chaque tour de boucle
                while True:
                    if duo_jobs.empty:
                        break

                    best = None
                    sample = duo_jobs.head(int(duo_pool)) if len(duo_jobs) > int(duo_pool) else duo_jobs

                    for jidx, job in sample.iterrows():
                        # Déduplication inter-jours DUO
                        if normalize_base_job_id(str(job["job_id"])) in planned_base_ids:
                            continue
                        addr = job["address"]
                        job_min_total = int(job["job_minutes"])
                        job_min_each = int(math.ceil(job_min_total / 2.0))
                        need_block = int(job_min_each) + int(buffer_job)

                        near_techs = rank_techs_for_job(tech_names, pd.Series({"address": addr}), int(techs_near_job))
                        jlat, jlon = get_ll_for_address(addr)
                        jsec = classify_sector(jlat, jlon)

                        for i in range(len(near_techs)):
                            for k in range(i + 1, len(near_techs)):
                                t1 = near_techs[i]
                                t2 = near_techs[k]

                                if lock_tech.get(t1) or lock_tech.get(t2):
                                    continue
                                if jobs_count[t1] >= int(max_jobs_per_day) or jobs_count[t2] >= int(max_jobs_per_day):
                                    continue
                                if not sector_compatible(_tech_sector.get(t1, "UNK"), jsec):
                                    continue
                                if not sector_compatible(_tech_sector.get(t2, "UNK"), jsec):
                                    continue

                                duo_is_overtime = False
                                if int(job_min_each) > int(daily_onsite_cap):
                                    _t1lat, _t1lon = tech_ll_map.get(t1, (None, None))
                                    _duo_tmin = travel_min_estimate(cur_loc[t1], job["address"], _t1lat, _t1lon, jlat, jlon)
                                    _duo_tback = travel_min_estimate(job["address"], _home_map[t1], jlat, jlon, _t1lat, _t1lon)
                                    _duo_need = _duo_tmin + int(job_min_each) + int(buffer_job) + _duo_tback
                                    if _duo_need <= int(OT_ACTIVE_CAP):
                                        if jobs_count[t1] != 0 or jobs_count[t2] != 0:
                                            continue
                                        duo_is_overtime = True
                                    else:
                                        continue

                                _t1lat, _t1lon = tech_ll_map.get(t1, (None, None))
                                _t2lat, _t2lon = tech_ll_map.get(t2, (None, None))
                                t1_tr = travel_min_estimate(cur_loc[t1], addr, _t1lat, _t1lon, jlat, jlon)
                                t2_tr = travel_min_estimate(cur_loc[t2], addr, _t2lat, _t2lon, jlat, jlon)
                                start_m = max(used[t1] + int(t1_tr), used[t2] + int(t2_tr))
                                end_m = start_m + int(need_block)

                                t1_back = travel_min_estimate(addr, _home_map[t1], jlat, jlon, _t1lat, _t1lon)
                                t2_back = travel_min_estimate(addr, _home_map[t2], jlat, jlon, _t2lat, _t2lon)

                                if (end_m + int(t1_back) <= available) and (end_m + int(t2_back) <= available):
                                    score = (start_m, max(int(t1_tr), int(t2_tr)))
                                    if best is None or score < best[0]:
                                        best = (
                                            score, jidx, t1, t2,
                                            start_m, end_m,
                                            int(t1_tr), int(t2_tr),
                                            duo_is_overtime,
                                            int(job_min_each)
                                        )

                    if best is None:
                        break

                    _, jidx, t1, t2, start_m, end_m, t1_tr_est, t2_tr_est, duo_is_overtime, job_min_each = best
                    job = duo_jobs.loc[jidx]
                    # Recalculer trajets réels pour le booking final
                    t1_tr = travel_min_cached(cur_loc[t1], job["address"])
                    t2_tr = travel_min_cached(cur_loc[t2], job["address"])
                    if t1_tr >= 9999: t1_tr = t1_tr_est
                    if t2_tr >= 9999: t2_tr = t2_tr_est
                    start_m = max(used[t1] + int(t1_tr), used[t2] + int(t2_tr))
                    end_m = start_m + int(job_min_each) + int(buffer_job)

                    for tname, trv in [(t1, t1_tr), (t2, t2_tr)]:
                        jobs_count[tname] += 1
                        planned_rows.append({
                            "date": day.isoformat(),
                            "technicien": tname,
                            "sequence": jobs_count[tname],
                            "job_id": job["job_id"],
                            "cust": job.get("cust", ""),
                            "duo": "⚠️ DUO",
                            "ot": "",
                            "debut": mm_to_hhmm(int(start_m)),
                            "fin": mm_to_hhmm(int(end_m)),
                            "adresse": job["address"],
                            "travel_min": int(trv),
                            "job_min": int(job_min_each),
                            "buffer_min": int(buffer_job),
                            "techs_needed": int(job["techs_needed"]),
                            **_extra_fields_from_job(job),
                            "description": job["description"],
                        })
                        planned_base_ids.add(normalize_base_job_id(job["job_id"]))
                        used[tname] = int(end_m)
                        cur_loc[tname] = job["address"]

                    if bool(duo_is_overtime):
                        lock_tech[t1] = True
                        lock_tech[t2] = True

                    duo_jobs = duo_jobs[duo_jobs["job_id"] != job["job_id"]].copy()

            # ---- 2) SOLO greedy per tech ----
            if not solo_jobs.empty:
                # [CRITIQUE-3] booked_ids : set O(1) au lieu de .copy() en boucle
                booked_ids: set = set()

                made_progress = True
                while made_progress:
                    made_progress = False
                    if solo_jobs.empty:
                        break

                    # Trier les techs par proximité à leur meilleur job disponible
                    # → les techs géographiquement spécialisés passent en premier
                    _active_techs = [t for t in tech_names if not lock_tech.get(t, False) and jobs_count[t] < int(max_jobs_per_day)]
                    _sorted_techs = _sort_techs_by_proximity(_active_techs, solo_jobs)
                    _locked_techs = [t for t in tech_names if lock_tech.get(t, False) or jobs_count[t] >= int(max_jobs_per_day)]
                    _ordered_techs = _sorted_techs + _locked_techs

                    for t in _ordered_techs:
                        if solo_jobs.empty:
                            break
                        if lock_tech.get(t, False):
                            continue
                        if jobs_count[t] >= int(max_jobs_per_day):
                            continue

                        # Filtrer le DataFrame une seule fois si booked_ids non vide
                        if booked_ids:
                            solo_jobs = solo_jobs[~solo_jobs["job_id"].isin(booked_ids)]
                            booked_ids.clear()
                            if solo_jobs.empty:
                                break

                        best_idx = None
                        best_cost = None
                        best_tmin = None

                        sample = get_job_pool_for_tech(solo_jobs, t, int(solo_pool))

                        for idx, job in sample.iterrows():
                            # Déduplication inter-jours : skip si déjà planifié
                            if normalize_base_job_id(str(job["job_id"])) in planned_base_ids:
                                continue
                            jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                            if not sector_compatible(_tech_sector.get(t, "UNK"), jsec):
                                continue
                            tlat_f, tlon_f = tech_ll_map.get(t, (None, None))
                            tmin = travel_min_estimate(cur_loc[t], job["address"], tlat_f, tlon_f, jlat, jlon)
                            tback = travel_min_estimate(job["address"], _home_map[t], jlat, jlon, tlat_f, tlon_f)
                            need = int(tmin) + int(job["job_minutes"]) + int(buffer_job) + int(tback)
                            if need <= 0:
                                continue
                            if used[t] + need <= available:
                                if best_cost is None or int(tmin) < best_cost:
                                    best_idx = idx
                                    best_cost = int(tmin)
                                    best_tmin = int(tmin)

                        if best_idx is not None:
                            job = jobs.loc[best_idx] if best_idx in jobs.index else solo_jobs.loc[best_idx]
                            jobs_count[t] += 1
                            # Recalculer avec API pour avoir l'heure précise dans le planning
                            best_tmin_real = travel_min_cached(cur_loc[t], job["address"])
                            if best_tmin_real >= 9999:
                                best_tmin_real = best_tmin  # fallback haversine
                            start_m = used[t] + int(best_tmin_real)
                            end_m = start_m + int(job["job_minutes"]) + int(buffer_job)

                            planned_rows.append({
                                "date": day.isoformat(),
                                "technicien": t,
                                "sequence": jobs_count[t],
                                "job_id": job["job_id"],
                                "cust": job.get("cust", ""),
                                "duo": "",
                                "ot": "",
                                "debut": mm_to_hhmm(int(start_m)),
                                "fin": mm_to_hhmm(int(end_m)),
                                "adresse": job["address"],
                                "travel_min": int(best_tmin_real),
                                "job_min": int(job["job_minutes"]),
                                "buffer_min": int(buffer_job),
                                "techs_needed": int(job["techs_needed"]),
                                **_extra_fields_from_job(job),
                                "description": job["description"],
                            })
                            planned_base_ids.add(normalize_base_job_id(job["job_id"]))
                            used[t] = int(end_m)
                            cur_loc[t] = job["address"]
                            booked_ids.add(job["job_id"])
                            made_progress = True
                            # Prioritize same-customer jobs next iteration
                            _booked_cust = str(job.get("cust", ""))
                            if _booked_cust and not solo_jobs.empty and "cust" in solo_jobs.columns:
                                _same_cust = solo_jobs["cust"].astype(str) == _booked_cust
                                if _same_cust.any():
                                    solo_jobs = pd.concat([
                                        solo_jobs[_same_cust],
                                        solo_jobs[~_same_cust]
                                    ])
                            continue

                        # OT single-job day
                        if jobs_count[t] == 0:
                            best_ot_idx = None
                            best_ot_cost = None
                            best_ot_tmin = None

                            for idx, job in sample.iterrows():
                                if normalize_base_job_id(str(job["job_id"])) in planned_base_ids:
                                    continue
                                jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                                if not sector_compatible(_tech_sector.get(t, "UNK"), jsec):
                                    continue
                                tlat_f, tlon_f = tech_ll_map.get(t, (None, None))
                                tmin = travel_min_estimate(cur_loc[t], job["address"], tlat_f, tlon_f, jlat, jlon)
                                tback = travel_min_estimate(job["address"], _home_map[t], jlat, jlon, tlat_f, tlon_f)
                                need = int(tmin) + int(job["job_minutes"]) + int(buffer_job) + int(tback)
                                if need <= OT_ACTIVE_CAP:
                                    if best_ot_cost is None or int(tmin) < best_ot_cost:
                                        best_ot_idx = idx
                                        best_ot_cost = int(tmin)
                                        best_ot_tmin = int(tmin)

                            if best_ot_idx is not None:
                                job = jobs.loc[best_ot_idx] if best_ot_idx in jobs.index else solo_jobs.loc[best_ot_idx]
                                jobs_count[t] += 1
                                start_m = used[t] + int(best_ot_tmin)
                                end_m = start_m + int(job["job_minutes"]) + int(buffer_job)

                                planned_rows.append({
                                    "date": day.isoformat(),
                                    "technicien": t,
                                    "sequence": jobs_count[t],
                                    "job_id": job["job_id"],
                                    "cust": job.get("cust", ""),
                                    "duo": "",
                                    "ot": "🟥 OT",
                                    "debut": mm_to_hhmm(int(start_m)),
                                    "fin": mm_to_hhmm(int(end_m)),
                                    "adresse": job["address"],
                                    "travel_min": int(best_ot_tmin),
                                    "job_min": int(job["job_minutes"]),
                                    "buffer_min": int(buffer_job),
                                    "techs_needed": int(job["techs_needed"]),
                                    **_extra_fields_from_job(job),
                                    "description": job["description"],
                                })
                                planned_base_ids.add(normalize_base_job_id(job["job_id"]))
                                used[t] = int(end_m)
                                cur_loc[t] = job["address"]
                                booked_ids.add(job["job_id"])
                                lock_tech[t] = True
                                made_progress = True
                                continue

                        # Split long jobs
                        best_long_idx = None
                        best_long_cost = None
                        best_long_is_overtime = False

                        for idx, job in sample.iterrows():
                            jm = int(job["job_minutes"])
                            if jm <= int(daily_onsite_cap):
                                continue
                            if t in carryover_by_tech:
                                continue
                            if normalize_base_job_id(str(job["job_id"])) in planned_base_ids:
                                continue
                            jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                            if not sector_compatible(_tech_sector.get(t, "UNK"), jsec):
                                continue

                            addr = job["address"]
                            tlat_f, tlon_f = tech_ll_map.get(t, (None, None))
                            tmin = travel_min_estimate(cur_loc[t], addr, tlat_f, tlon_f, jlat, jlon)
                            tback = travel_min_estimate(addr, _home_map[t], jlat, jlon, tlat_f, tlon_f)

                            # Décision OT-en-une-journée vs split :
                            # Si trajet + job + buffer + retour <= 14h → OT en une journée
                            # Sinon → split sur plusieurs jours
                            full_need = int(tmin) + int(jm) + int(buffer_job) + int(tback)
                            is_overtime_candidate = (
                                jobs_count[t] == 0
                                and full_need <= int(OT_ACTIVE_CAP)
                            )

                            max_onsite_today = int(available) - int(used[t]) - int(tmin) - int(buffer_job) - int(tback)
                            if max_onsite_today <= 0:
                                continue
                            if jobs_count[t] > 0 and int(max_onsite_today) < int(MIN_ONSITE_CHUNK_MIN):
                                continue

                            onsite_today_candidate = choose_onsite_no_crumbs(jm, max_onsite_today, MIN_ONSITE_CHUNK_MIN)
                            if onsite_today_candidate <= 0:
                                continue
                            # Si OT candidat, on book en entier plus bas — pas de split partiel
                            if is_overtime_candidate:
                                pass  # sera booké en entier dans le bloc best_long_is_overtime

                            if best_long_cost is None or int(tmin) < best_long_cost:
                                best_long_idx = idx
                                best_long_cost = int(tmin)
                                best_long_is_overtime = bool(is_overtime_candidate)

                        if best_long_idx is None:
                            continue

                        job = jobs.loc[best_long_idx] if best_long_idx in jobs.index else solo_jobs.loc[best_long_idx]
                        base_job_id = str(job["job_id"])
                        jm_total = int(job["job_minutes"])

                        if bool(best_long_is_overtime):
                            tmin = travel_min_cached(cur_loc[t], job["address"])
                            start_m = int(used[t]) + int(tmin)
                            end_m = start_m + int(jm_total) + int(buffer_job)

                            jobs_count[t] += 1
                            planned_rows.append({
                                "date": day.isoformat(),
                                "technicien": t,
                                "sequence": jobs_count[t],
                                "job_id": job["job_id"],
                                "cust": job.get("cust", ""),
                                "duo": "",
                                "ot": "🟥 OT",
                                "debut": mm_to_hhmm(int(start_m)),
                                "fin": mm_to_hhmm(int(end_m)),
                                "adresse": job["address"],
                                "travel_min": int(tmin),
                                "job_min": int(jm_total),
                                "buffer_min": int(buffer_job),
                                "techs_needed": int(job.get("techs_needed", 1)),
                                **_extra_fields_from_job(job),
                                "description": job["description"],
                            })
                            planned_base_ids.add(normalize_base_job_id(base_job_id))
                            used[t] = int(end_m)
                            cur_loc[t] = job["address"]
                            booked_ids.add(job["job_id"])
                            lock_tech[t] = True
                            made_progress = True
                            continue

                        total_parts_guess = compute_total_parts(int(job["job_minutes"]), int(daily_onsite_cap))
                        carryover_by_tech[t] = {
                            "base_job_id": base_job_id,
                            "cust": job.get("cust", ""),
                            "address": job["address"],
                            "description": job["description"],
                            "techs_needed": int(job.get("techs_needed", 1)),
                            "total_parts": int(total_parts_guess),
                            "part_idx_next": 1,
                            "remaining_job_min": int(job["job_minutes"]),
                            "last_inspection": job.get("last_inspection", ""),
                            "difference": job.get("difference", ""),
                            "unit": job.get("unit", ""),
                            "serial_number": job.get("serial_number", ""),
                        }
                        booked_ids.add(job["job_id"])

                        _book_split_part_for_tech(day, t, used, cur_loc, jobs_count, lock_tech, carryover_by_tech[t])
                        if t in carryover_by_tech and int(carryover_by_tech[t]["remaining_job_min"]) <= 0:
                            del carryover_by_tech[t]

                        made_progress = True

                # Appliquer les booked_ids restants au DataFrame
                if booked_ids:
                    solo_jobs = solo_jobs[~solo_jobs["job_id"].isin(booked_ids)]

            # RETURN_HOME
            for t in tech_names:
                if jobs_count[t] > 0:
                    tback = travel_min_cached(cur_loc[t], _home_map[t])
                    planned_rows.append({
                        "date": day.isoformat(),
                        "technicien": t,
                        "sequence": jobs_count[t] + 1,
                        "job_id": "RETURN_HOME",
                        "cust": "",
                        "duo": "",
                        "ot": "",
                        "debut": mm_to_hhmm(int(used[t])),
                        "fin": mm_to_hhmm(int(used[t]) + int(tback)),
                        "adresse": _home_map[t],
                        "travel_min": int(tback),
                        "job_min": 0,
                        "buffer_min": 0,
                        "techs_needed": 1,
                        "last_inspection": "",
                        "difference": "",
                        "unit": "",
                        "serial_number": "",
                        "description": "🏠 Retour domicile (estimé)",
                    })
                    used[t] = int(used[t]) + int(tback)
                    cur_loc[t] = _home_map[t]

            if progress is not None:
                progress.progress(int(((di + 1) / total_steps) * 100))
            if progress_text is not None:
                _t_day = round(time.time() - _t_day_start, 1)
                progress_text.write(
                    f"Planification… {di+1}/{len(month_days)} jour(s) — "
                    f"jour actuel: {_t_day}s | "
                    f"API calls: {st.session_state.get('p2_api_calls',0)} | "
                    f"cache hits: {st.session_state.get('p2_cache_hits',0)}"
                )

        _t_post = time.time()
        _t_planning_done = _t_post
        if progress_text is not None:
            progress_text.write(f"✅ Planification terminée — post-traitement en cours…")

        # Réinjecter carryovers non terminés
        carryover_rows = []
        for t, stt in carryover_by_tech.items():
            rem = int(stt.get("remaining_job_min", 0))
            if rem > 0:
                carryover_rows.append({
                    "job_id": str(stt.get("base_job_id", "")),
                    "cust": str(stt.get("cust", "")),
                    "address": str(stt.get("address", "")),
                    "description": str(stt.get("description", "")),
                    "job_minutes": rem,
                    "techs_needed": int(stt.get("techs_needed", 1)),
                    "postal": extract_postal(str(stt.get("address", ""))),
                    "last_inspection": str(stt.get("last_inspection", "")),
                    "difference": str(stt.get("difference", "")),
                    "unit": str(stt.get("unit", "")),
                    "serial_number": str(stt.get("serial_number", "")),
                })

        remaining_out = pd.concat([duo_jobs, solo_jobs, hard_jobs], ignore_index=True)
        if carryover_rows:
            remaining_out = pd.concat([remaining_out, pd.DataFrame(carryover_rows)], ignore_index=True)

        _t_concat = round(time.time() - _t_post, 1)
        if progress_text is not None:
            progress_text.write(f"⏱️ Concat remaining: {_t_concat}s — {len(remaining_out)} jobs restants")

        # Flag OT-impossible
        # Ignoré si le cache n'est pas chaud — paires domicile→job souvent absentes
        # Timeout 30s pour éviter de bloquer après la planification
        if not remaining_out.empty:
            remaining_out = remaining_out.copy()
            remaining_out["ot_impossible"] = False
            OT_IMPOSSIBLE_TOP_TECHS = 4
            _best_need_cache = {}
            _ot_flag_start = time.time()
            _ot_flag_timeout = False

            for i, r in remaining_out.iterrows():
                # Timeout 30s — l'indicateur OT-impossible est secondaire
                if time.time() - _ot_flag_start > 10:
                    _ot_flag_timeout = True
                    break
                try:
                    addr = str(r.get("address", "")).strip()
                    jm = int(r.get("job_minutes", 0))
                    if not addr or jm <= 0:
                        continue

                    ck = (addr, jm, int(buffer_job), OT_ACTIVE_CAP, OT_IMPOSSIBLE_TOP_TECHS)
                    if ck in _best_need_cache:
                        best_need = _best_need_cache[ck]
                    else:
                        jlat, jlon = get_ll_for_address(addr)
                        jsec = classify_sector(jlat, jlon)
                        scored = []
                        for t in tech_names:
                            tsec = _tech_sector.get(t, "UNK")
                            if not sector_compatible(tsec, jsec):
                                continue
                            tlat, tlon = tech_ll_map.get(t, (None, None))
                            d = haversine_km(tlat, tlon, jlat, jlon)
                            scored.append((d, t))

                        if not scored:
                            candidates = list(tech_names)[:OT_IMPOSSIBLE_TOP_TECHS]
                        else:
                            scored.sort(key=lambda x: x[0])
                            candidates = [t for (_d, t) in scored[:OT_IMPOSSIBLE_TOP_TECHS]]

                        best_need = None
                        for t in candidates:
                            # Utiliser seulement le cache SQLite — pas d'appel API si absent
                            k_fwd = _key(_home_map[t], addr, use_traffic)
                            k_bck = _key(addr, _home_map[t], use_traffic)
                            conn = _get_db()
                            cur = conn.cursor()
                            now = int(time.time())
                            min_ts = now - int(cache_days) * 86400
                            cur.execute("SELECT minutes FROM travel WHERE k=? AND ts>=?", (k_fwd, min_ts))
                            row_fwd = cur.fetchone()
                            cur.execute("SELECT minutes FROM travel WHERE k=? AND ts>=?", (k_bck, min_ts))
                            row_bck = cur.fetchone()
                            # Si une des deux paires manque dans le cache → ignorer ce job
                            if not row_fwd or not row_bck:
                                continue
                            need = int(row_fwd[0]) + jm + int(buffer_job) + int(row_bck[0])
                            if best_need is None or need < best_need:
                                best_need = need
                        _best_need_cache[ck] = best_need

                    if best_need is not None and best_need > OT_ACTIVE_CAP:
                        remaining_out.at[i, "ot_impossible"] = True
                except Exception:
                    pass

        _t_ot = round(time.time() - _t_post, 1)
        if progress_text is not None:
            progress_text.write(f"⏱️ Flag OT-impossible: {_t_ot}s")

        # ── BACKFILL PASS ──────────────────────────────────────────────────────
        # Deuxième passage sur les jobs restants — essaie de les insérer sur
        # n'importe quel jour encore disponible pour les techs compatibles.
        # Corrige le problème greedy myope : le scheduler s'arrêtait dès qu'un
        # tour complet ne produisait rien, même avec 1000h de capacité libre.
        # ───────────────────────────────────────────────────────────────────────
        if not remaining_out.empty and month_days:
            if progress_text is not None:
                progress_text.write(f"🔄 Backfill pass — {len(remaining_out)} jobs restants…")

            # Reconstruire l'utilisation actuelle par (jour, tech)
            _used_by_day_tech: Dict[Tuple[str,str], int] = {}
            _loc_by_day_tech: Dict[Tuple[str,str], str] = {}
            _count_by_day_tech: Dict[Tuple[str,str], int] = {}

            for row in planned_rows:
                dk = str(row.get("date","")).strip()
                tk = str(row.get("technicien","")).strip()
                if not dk or not tk:
                    continue
                if _is_return_home(row.get("job_id","")):
                    continue
                fin_mm = _hhmm_to_mm(str(row.get("fin","00:00")))
                key = (dk, tk)
                _used_by_day_tech[key] = max(_used_by_day_tech.get(key, 0), int(fin_mm))
                _loc_by_day_tech[key] = str(row.get("adresse", _home_map.get(tk,"")))
                _count_by_day_tech[key] = _count_by_day_tech.get(key, 0) + 1

            backfill_booked = set()
            backfill_rows = []

            # Trier les jobs restants par contrainte croissante (moins de techs = plus urgent)
            remaining_solo = remaining_out[
                (remaining_out["techs_needed"] <= 1) &
                (remaining_out.get("ot_impossible", pd.Series([False]*len(remaining_out))).fillna(False) == False)
            ].copy()

            for _, jrow in remaining_solo.iterrows():
                jid = str(jrow.get("job_id",""))
                # Vérifier que ce job n'est pas déjà planifié (greedy principal ou backfill)
                if jid in backfill_booked:
                    continue
                if normalize_base_job_id(jid) in planned_base_ids:
                    continue
                addr = str(jrow.get("address","")).strip()
                jm = int(jrow.get("job_minutes", 0))
                if not addr or jm <= 0:
                    continue

                jlat, jlon = get_ll_for_address(addr)
                jsec = classify_sector(jlat, jlon)

                booked = False
                for day in month_days:
                    if booked:
                        break
                    dk = day.isoformat()
                    for t in tech_names:
                        tsec = _tech_sector.get(t, "UNK")
                        if not sector_compatible(tsec, jsec):
                            continue

                        key = (dk, t)
                        cur_used = _used_by_day_tech.get(key, 0)
                        cur_count = _count_by_day_tech.get(key, 0)
                        home_addr = _home_map.get(t, "")

                        if cur_count >= int(max_jobs_per_day):
                            continue

                        # Utiliser estimate pour l'évaluation backfill (0 appel API)
                        # travel_min_cached est appelé uniquement au booking final
                        _t_blat, _t_blon = tech_ll_map.get(t, (None, None))
                        _j_blat = jobs.at[jidx, "job_lat"] if jidx in jobs.index and "job_lat" in jobs.columns else None
                        _j_blon = jobs.at[jidx, "job_lon"] if jidx in jobs.index and "job_lon" in jobs.columns else None
                        tmin = travel_min_estimate(home_addr, addr, _t_blat, _t_blon, _j_blat, _j_blon)
                        tback = travel_min_estimate(addr, home_addr, _j_blat, _j_blon, _t_blat, _t_blon)
                        need = int(tmin) + int(jm) + int(buffer_job) + int(tback)

                        # Rentre dans la journée normale?
                        fits_normal = (cur_used + need) <= int(available)
                        # Rentre en OT (max 14h)?
                        fits_ot = (cur_used == 0) and (need <= int(OT_ACTIVE_CAP))

                        if not fits_normal and not fits_ot:
                            continue

                        # Booker ce job
                        # Recalculer avec API pour l'heure précise au booking
                        tmin_real = travel_min_cached(home_addr, addr)
                        if tmin_real >= 9999: tmin_real = tmin
                        tback_real = travel_min_cached(addr, home_addr)
                        if tback_real >= 9999: tback_real = tback
                        start_m = cur_used + int(tmin_real)
                        end_m = start_m + int(jm) + int(buffer_job)
                        ot_flag = "🟥 OT" if fits_ot and not fits_normal else ""

                        new_row = {
                            "date": dk,
                            "technicien": t,
                            "sequence": cur_count + 1,
                            "job_id": jid,
                            "cust": str(jrow.get("cust","")),
                            "duo": "",
                            "ot": ot_flag,
                            "debut": mm_to_hhmm(int(start_m)),
                            "fin": mm_to_hhmm(int(end_m)),
                            "adresse": addr,
                            "travel_min": int(tmin),
                            "job_min": int(jm),
                            "buffer_min": int(buffer_job),
                            "techs_needed": int(jrow.get("techs_needed", 1)),
                            "last_inspection": str(jrow.get("last_inspection","")),
                            "difference": str(jrow.get("difference","")),
                            "unit": str(jrow.get("unit","")),
                            "serial_number": str(jrow.get("serial_number","")),
                            "description": str(jrow.get("description","")),
                        }
                        backfill_rows.append(new_row)
                        planned_base_ids.add(normalize_base_job_id(jid))
                        backfill_booked.add(jid)

                        # Mettre à jour l'état du jour/tech
                        _used_by_day_tech[key] = int(end_m)
                        _loc_by_day_tech[key] = addr
                        _count_by_day_tech[key] = cur_count + 1
                        booked = True
                        break

            if backfill_rows:
                planned_rows.extend(backfill_rows)
                # Retirer les jobs backfillés de remaining_out
                remaining_out = remaining_out[
                    ~remaining_out["job_id"].astype(str).isin(backfill_booked)
                ].copy()
                if progress_text is not None:
                    progress_text.write(f"✅ Backfill: {len(backfill_rows)} jobs ajoutés")

        # ── FIN BACKFILL PASS ──────────────────────────────────────────────────

        # Safety net : réinjecter jobs bookables manquants
        try:
            all_bookable_ids = set(jobs_in["job_id"].astype(str).apply(normalize_base_job_id).tolist())
            remaining_ids = set()
            if (remaining_out is not None) and (not remaining_out.empty) and ("job_id" in remaining_out.columns):
                remaining_ids = set(remaining_out["job_id"].astype(str).apply(normalize_base_job_id).tolist())

            covered = planned_base_ids.union(remaining_ids)
            missing = sorted(list(all_bookable_ids - covered))

            if missing:
                if "ot_impossible" not in remaining_out.columns:
                    remaining_out["ot_impossible"] = False
                if "missing_reinjected" not in remaining_out.columns:
                    remaining_out["missing_reinjected"] = False

                missing_rows = []
                for mid in missing:
                    src = jobs_in[jobs_in["job_id"].astype(str) == str(mid)]
                    if src.empty:
                        continue
                    rr = src.iloc[0].to_dict()
                    rr["missing_reinjected"] = True
                    rr["ot_impossible"] = False
                    missing_rows.append(rr)

                if missing_rows:
                    remaining_out = pd.concat([remaining_out, pd.DataFrame(missing_rows)], ignore_index=True)
        except Exception:
            pass

        _t_safety = round(time.time() - _t_post, 1)
        _timing_summary = {
            "concat_s": _t_concat if "_t_concat" in dir() else "?",
            "ot_flag_s": _t_ot if "_t_ot" in dir() else "?",
            "safety_s": _t_safety,
            "total_post_s": round(time.time() - _t_post, 1),
        }
        if progress_text is not None:
            progress_text.write(f"⏱️ Safety net: {_t_safety}s — returning…")

        success = remaining_out.empty and (len(carryover_rows) == 0)
        return {"success": bool(success), "rows": planned_rows, "remaining": remaining_out,
                "timing": _timing_summary}


    # ════════════════════════════════════════════════════════════════
    # Repair pass + helpers
    # ════════════════════════════════════════════════════════════════
    def _is_return_home(job_id: str) -> bool:
        return str(job_id).strip().upper() == "RETURN_HOME"

    def _is_split_part(job_id: str) -> bool:
        return "(PART" in str(job_id).upper()

    def _is_duo_row(row: dict) -> bool:
        return str(row.get("duo", "")).strip() != ""

    def _is_ot_row(row: dict) -> bool:
        return str(row.get("ot", "")).strip() != ""

    def _hhmm_to_mm(hhmm: str) -> int:
        try:
            s = str(hhmm or "").strip()
            if not s:
                return 0
            h, m = s.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return 0

    def optimize_route_ortools(tech: str, jobs_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        PRIORITÉ 2 — OR-Tools route optimization.
        Retourne jobs_list réordonné pour minimiser le trajet total.
        Utilise uniquement le cache SQLite — zéro appel API si matrice précalculée.

        Si ortools n'est pas installé, retourne l'ordre original (greedy).
        Fallback automatique si la matrice contient des valeurs 9999 (paires inconnues).
        """
        if not ORTOOLS_AVAILABLE or len(jobs_list) <= 1:
            return jobs_list

        home_addr = home_map[tech]
        # Nœuds : 0=domicile départ, 1..n=jobs, n+1=domicile retour
        all_locs = [home_addr] + [jb["adresse"] for jb in jobs_list] + [home_addr]
        n = len(all_locs)

        # Construire la matrice depuis SQLite (tout doit être en cache)
        matrix = []
        has_missing = False
        # Coordonnées pour haversine dans la matrice OR-Tools
        _locs_ll = []
        for _loc in all_locs:
            _ll = get_ll_for_address(_loc) if _loc else (None, None)
            _locs_ll.append(_ll)

        for i in range(n):
            row_m = []
            for j in range(n):
                if i == j:
                    row_m.append(0)
                else:
                    # Vérifier le cache SQLite d'abord, puis haversine
                    _ilat, _ilon = _locs_ll[i]
                    _jlat, _jlon = _locs_ll[j]
                    v = travel_min_estimate(all_locs[i], all_locs[j], _ilat, _ilon, _jlat, _jlon)
                    if v >= 9999:
                        has_missing = True
                    row_m.append(int(v))
            matrix.append(row_m)

        # Si des paires sont manquantes → fallback greedy (sera comblé au prochain prefetch)
        if has_missing:
            return jobs_list

        try:
            manager = pywrapcp.RoutingIndexManager(n, 1, [0], [n - 1])
            routing = pywrapcp.RoutingModel(manager)

            def distance_callback(from_idx, to_idx):
                return matrix[manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]

            transit_idx = routing.RegisterTransitCallback(distance_callback)
            routing.SetArcCostEvaluatorOfAllVehicles(transit_idx)

            params = pywrapcp.DefaultRoutingSearchParameters()
            params.first_solution_strategy = (
                routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
            )
            params.local_search_metaheuristic = (
                routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
            )
            params.time_limit.seconds = 2  # max 2s par technicien

            solution = routing.SolveWithParameters(params)
            if not solution:
                return jobs_list

            # Extraire l'ordre optimal
            order = []
            index = routing.Start(0)
            while not routing.IsEnd(index):
                node = manager.IndexToNode(index)
                if 0 < node < n - 1:  # ignorer nœud départ (0) et arrivée (n-1)
                    order.append(node - 1)  # convertir en index jobs_list
                index = solution.Value(routing.NextVar(index))

            if len(order) != len(jobs_list):
                return jobs_list

            return [jobs_list[i] for i in order]

        except Exception:
            return jobs_list

    def rebuild_day_greedy(day: str, tech: str, jobs_list: List[Dict[str, Any]],
                           day_available_min: int, buffer_job: int, max_jobs_per_day: int,
                           _travel_fn=None):
        """
        _travel_fn : fonction (a, b) -> int optionnelle.
        Si fournie (ex: mini-cache local du repair pass), utilisée à la place
        de travel_min_cached pour éviter les appels API redondants.
        """
        home_addr = home_map[tech]
        cur = home_addr
        used = 0
        seq = 0
        _tfn = _travel_fn if _travel_fn is not None else travel_min_estimate
        # OR-Tools : seulement si appelé directement (pas depuis repair)
        # Dans repair, _travel_fn est fourni — on skip OR-Tools pour éviter
        # de reconstruire une matrice complète à chaque rebuild testé
        if _travel_fn is None:
            optimized = optimize_route_ortools(tech, list(jobs_list))
        else:
            optimized = list(jobs_list)
        remaining = list(optimized)
        out_rows: List[Dict[str, Any]] = []

        if len(remaining) > int(max_jobs_per_day):
            return None

        while remaining:
            best_i = None
            best_tmin = None
            for i, jb in enumerate(remaining):
                tmin = _tfn(cur, jb["adresse"])
                if best_tmin is None or int(tmin) < int(best_tmin):
                    best_tmin = int(tmin)
                    best_i = i

            jb = remaining.pop(best_i)
            tmin = int(best_tmin)
            tback = int(_tfn(jb["adresse"], home_addr))
            need = tmin + int(jb["job_min"]) + int(buffer_job) + tback

            if used + need > int(day_available_min):
                return None

            seq += 1
            start_m = used + tmin
            end_m = start_m + int(jb["job_min"]) + int(buffer_job)

            out_rows.append({
                "date": day,
                "technicien": tech,
                "sequence": seq,
                "job_id": jb["job_id"],
                "cust": jb.get("cust", ""),
                "duo": "",
                "ot": jb.get("ot", ""),
                "debut": mm_to_hhmm(int(start_m)),
                "fin": mm_to_hhmm(int(end_m)),
                "adresse": jb["adresse"],
                "travel_min": int(tmin),
                "job_min": int(jb["job_min"]),
                "buffer_min": int(buffer_job),
                "techs_needed": int(jb.get("techs_needed", 1)),
                "unit": jb.get("unit", ""),
                "serial_number": jb.get("serial_number", ""),
                "difference": jb.get("difference", ""),
                "last_inspection": jb.get("last_inspection", ""),
                "description": jb.get("description", ""),
            })

            used = int(end_m)
            cur = jb["adresse"]

            if seq >= int(max_jobs_per_day) and remaining:
                return None

        if seq > 0:
            tback = int(_tfn(cur, home_addr))
            out_rows.append({
                "date": day, "technicien": tech, "sequence": seq + 1,
                "job_id": "RETURN_HOME", "cust": "", "duo": "", "ot": "",
                "debut": mm_to_hhmm(int(used)),
                "fin": mm_to_hhmm(int(used) + int(tback)),
                "adresse": home_addr, "travel_min": int(tback), "job_min": 0,
                "buffer_min": 0, "techs_needed": 1, "unit": "", "serial_number": "",
                "difference": "", "last_inspection": "", "description": "🏠 Retour domicile (estimé)",
            })
        return out_rows

    def rebuild_day_in_order(day, tech, jobs_list, day_available_min, buffer_job, max_jobs_per_day,
                              _travel_fn=None):
        home_addr = home_map[tech]
        cur = home_addr
        used = 0
        seq = 0
        out = []
        _tfn2 = _travel_fn if _travel_fn is not None else travel_min_estimate

        if len(jobs_list) > int(max_jobs_per_day):
            return None

        for jb in jobs_list:
            tmin = int(_tfn2(cur, jb["adresse"]))
            tback = int(_tfn2(jb["adresse"], home_addr))
            need = tmin + int(jb["job_min"]) + int(buffer_job) + tback

            if used + need > int(day_available_min):
                return None

            seq += 1
            start_m = used + tmin
            end_m = start_m + int(jb["job_min"]) + int(buffer_job)

            out.append({
                "date": day, "technicien": tech, "sequence": seq,
                "job_id": jb["job_id"], "cust": jb.get("cust", ""),
                "duo": "", "ot": jb.get("ot", ""),
                "debut": mm_to_hhmm(int(start_m)),
                "fin": mm_to_hhmm(int(end_m)),
                "adresse": jb["adresse"], "travel_min": int(tmin),
                "job_min": int(jb["job_min"]), "buffer_min": int(buffer_job),
                "techs_needed": int(jb.get("techs_needed", 1)),
                "unit": jb.get("unit", ""), "serial_number": jb.get("serial_number", ""),
                "difference": jb.get("difference", ""), "last_inspection": jb.get("last_inspection", ""),
                "description": jb.get("description", ""),
            })

            used = int(end_m)
            cur = jb["adresse"]

            if seq >= int(max_jobs_per_day) and (len(out) < len(jobs_list)):
                return None

        if seq > 0:
            tback = int(_tfn2(cur, home_addr))
            out.append({
                "date": day, "technicien": tech, "sequence": seq + 1,
                "job_id": "RETURN_HOME", "cust": "", "duo": "", "ot": "",
                "debut": mm_to_hhmm(int(used)),
                "fin": mm_to_hhmm(int(used) + int(tback)),
                "adresse": home_addr, "travel_min": int(tback), "job_min": 0,
                "buffer_min": 0, "techs_needed": 1, "unit": "", "serial_number": "",
                "difference": "", "last_inspection": "", "description": "🏠 Retour domicile (estimé)",
            })
        return out

    def repair_month_plan(planned_rows, tech_names, day_available_min, buffer_job,
                          max_jobs_per_day, threshold_travel_min=50, candidate_techs_top_n=4, max_moves=25,
                          timeout_seconds=30):
        if not planned_rows:
            return planned_rows, {"moves": 0, "attempts": 0, "improved": 0}

        # [ÉLEVÉ-3] Mini-cache local travel pour éviter appels redondants dans repair
        _travel_local: Dict[Tuple[str,str], int] = {}
        def _travel(a: str, b: str) -> int:
            k = (a, b)
            if k not in _travel_local:
                # Utiliser estimate pour le repair (évite appels API massifs)
                # Le repair évalue des dizaines de swaps → estimation suffit
                _travel_local[k] = travel_min_estimate(a, b)
            return _travel_local[k]

        locked, movable = [], []
        original_rows_by_day_tech: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        original_return_rows: Dict[Tuple[str, str], Dict[str, Any]] = {}

        for r in planned_rows:
            day_k = str(r.get("date", "")).strip()
            tech_k = str(r.get("technicien", "")).strip()

            if _is_return_home(r.get("job_id", "")):
                original_return_rows[(day_k, tech_k)] = r

            if _is_return_home(r.get("job_id", "")):
                locked.append(r)
            elif _is_duo_row(r) or _is_split_part(r.get("job_id", "")) or _is_ot_row(r):
                locked.append(r)
            else:
                movable.append(r)
                original_rows_by_day_tech.setdefault((day_k, tech_k), []).append(r)

        by_day_tech: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        for r in movable:
            k = (str(r.get("date", "")).strip(), str(r.get("technicien", "")).strip())
            by_day_tech.setdefault(k, []).append({
                "job_id": r["job_id"],
                "cust": r.get("cust", ""),
                "ot": r.get("ot", ""),
                "adresse": r["adresse"],
                "job_min": int(r.get("job_min", 0)),
                "techs_needed": int(r.get("techs_needed", 1)),
                "description": r.get("description", ""),
                "unit": r.get("unit", ""),
                "serial_number": r.get("serial_number", ""),
                "difference": r.get("difference", ""),
                "last_inspection": r.get("last_inspection", ""),
                "travel_min_orig": int(r.get("travel_min", 0)),
            })

        candidates = []
        for (day, tech), lst in by_day_tech.items():
            for jb in lst:
                candidates.append((int(jb["travel_min_orig"]), day, tech, jb))
        candidates.sort(key=lambda x: x[0], reverse=True)

        stats = {"moves": 0, "attempts": 0, "improved": 0, "timeout": False}
        tech_list = list(tech_names)
        _repair_start = time.time()
        _modified_pairs: set = set()  # paires (day, tech) modifiées — seules à rebuilder

        for (tmin_orig, day, src_tech, jb) in candidates:
            if time.time() - _repair_start > float(timeout_seconds):
                stats["timeout"] = True
                break
            if stats["moves"] >= int(max_moves):
                break
            if int(tmin_orig) <= int(threshold_travel_min):
                break

            stats["attempts"] += 1

            job_row = pd.Series({"address": jb["adresse"], "job_id": jb["job_id"], "job_minutes": int(jb["job_min"])})
            near_techs = rank_techs_for_job(tech_list, job_row, int(candidate_techs_top_n))
            near_techs = [t for t in near_techs if t != src_tech]
            if not near_techs:
                continue

            best_move = None

            for dst_tech in near_techs:
                src_key = (day, src_tech)
                dst_key = (day, dst_tech)

                src_list = list(by_day_tech.get(src_key, []))
                dst_list = list(by_day_tech.get(dst_key, []))

                src_list2 = [x for x in src_list if str(x["job_id"]) != str(jb["job_id"])]
                dst_list2 = dst_list + [jb]

                if len(dst_list2) > int(max_jobs_per_day):
                    continue

                rebuilt_src = rebuild_day_greedy(day, src_tech, src_list2, day_available_min, buffer_job, max_jobs_per_day, _travel_fn=_travel)
                rebuilt_dst = rebuild_day_greedy(day, dst_tech, dst_list2, day_available_min, buffer_job, max_jobs_per_day, _travel_fn=_travel)
                if rebuilt_src is None or rebuilt_dst is None:
                    continue

                new_tmin = None
                for rr in rebuilt_dst:
                    if str(rr.get("job_id")) == str(jb["job_id"]):
                        new_tmin = int(rr.get("travel_min", 9999))
                        break
                if new_tmin is None:
                    continue

                delta = int(tmin_orig) - int(new_tmin)
                if delta <= 0:
                    continue

                if best_move is None or delta > best_move[0]:
                    best_move = (delta, dst_tech, rebuilt_src, rebuilt_dst, src_list2, dst_list2)

            if best_move is None:
                continue

            delta, dst_tech, rebuilt_src, rebuilt_dst, src_list2, dst_list2 = best_move
            by_day_tech[(day, src_tech)] = src_list2
            by_day_tech[(day, dst_tech)] = dst_list2
            _modified_pairs.add((day, src_tech))
            _modified_pairs.add((day, dst_tech))
            stats["moves"] += 1
            stats["improved"] += 1

        final_rows: List[Dict[str, Any]] = []

        for r in locked:
            if _is_return_home(r.get("job_id", "")):
                continue
            final_rows.append(r)

        for (day, tech), lst in by_day_tech.items():
            if (day, tech) in _modified_pairs:
                # Seulement rebuilder les paires qui ont été modifiées
                rebuilt = rebuild_day_greedy(day, tech, lst, day_available_min, buffer_job, max_jobs_per_day, _travel_fn=_travel)
                if rebuilt is not None:
                    final_rows.extend(rebuilt)
                else:
                    orig = original_rows_by_day_tech.get((day, tech), [])
                    if orig:
                        orig_sorted = sorted(orig, key=lambda x: int(x.get("sequence", 0)))
                        rebuilt2 = rebuild_day_in_order(day, tech, orig_sorted, day_available_min, buffer_job, max_jobs_per_day, _travel_fn=_travel)
                        if rebuilt2 is not None:
                            final_rows.extend(rebuilt2)
                        else:
                            final_rows.extend(orig)
            else:
                # Paire non modifiée — réutiliser les rows originaux directement
                orig = original_rows_by_day_tech.get((day, tech), [])
                final_rows.extend(orig)

        final_rows = [r for r in final_rows if not _is_return_home(r.get("job_id", ""))]

        groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        for r in final_rows:
            day_k = str(r.get("date", "")).strip()
            tech_k = str(r.get("technicien", "")).strip()
            if not day_k or not tech_k:
                continue
            groups.setdefault((day_k, tech_k), []).append(r)

        all_pairs = set(original_return_rows.keys()).union(set(groups.keys()))

        for (day_k, tech_k) in sorted(all_pairs):
            rows = groups.get((day_k, tech_k), [])
            real_rows = [x for x in rows if not _is_return_home(x.get("job_id", ""))]
            if not real_rows:
                continue

            last_row = max(real_rows, key=lambda x: _hhmm_to_mm(x.get("fin", "00:00")))

            home_addr = home_map.get(tech_k)
            if not home_addr:
                continue

            cur_addr = str(last_row.get("adresse", "")).strip() or home_addr
            used_end = _hhmm_to_mm(last_row.get("fin", "00:00"))
            tback = int(_travel(cur_addr, home_addr))

            final_rows.append({
                "date": day_k, "technicien": tech_k,
                "sequence": int(last_row.get("sequence", 0)) + 1,
                "job_id": "RETURN_HOME", "cust": "", "duo": "", "ot": "",
                "debut": mm_to_hhmm(int(used_end)),
                "fin": mm_to_hhmm(int(used_end) + int(tback)),
                "adresse": home_addr, "travel_min": int(tback), "job_min": 0,
                "buffer_min": 0, "techs_needed": 1, "unit": "", "serial_number": "",
                "difference": "", "last_inspection": "", "description": "🏠 Retour domicile (estimé)",
            })

        return final_rows, stats

    # [MOYEN-3] Integrity check — sets stockés dans session_state, pas recalculés

    def build_export_excel(df: pd.DataFrame, unplanned_df: pd.DataFrame = None) -> bytes:
        """
        Builds a styled Excel file from the planning DataFrame.
        - Sheet 1: planning with last_inspection split into 3 cols + color coding
        - Sheet 2 (optional): unplanned jobs
        - Description is truncated to first 4 upcoming service instances
        """
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from datetime import datetime, date
        import re as _re

        today = date.today()

        def parse_last_insp(val):
            s = str(val or "").strip()
            if not s or s == "nan":
                return "", "", ""
            parts = [p.strip() for p in s.split("//")]
            wo  = parts[0] if len(parts) > 0 else ""
            svc = parts[1] if len(parts) > 1 else ""
            dt  = parts[2] if len(parts) > 2 else ""
            return wo, svc, dt

        def date_color(date_str):
            """
            Retourne (bg_hex_8, fg_hex_6) selon la distance de la date par rapport à aujourd'hui.
            - diff < -180j (>6 mois passé)       → brown  + texte blanc
            - diff < 120j  (passé récent / <4mo) → gris
            - diff 120-150j (4-5 mois futur)     → pink
            - diff 150-180j (5-6 mois futur)     → orange
            - diff > 180j  (>6 mois futur)       → rien
            Les codes sont en format 8-char ARGB pour PatternFill (openpyxl).
            """
            s = str(date_str or "").strip()
            if not s:
                return None, None
            for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    dt = datetime.strptime(s, fmt).date()
                    diff = (dt - today).days
                    if diff < -180:   return "FF7B4A2D", "FFFFFF"  # >6mo passé → brown
                    elif diff < 120:  return "FFCCCCCC", "000000"  # passé récent ou <4mo → gris
                    elif diff < 150:  return "FFFFB6C1", "000000"  # 4-5mo futur → pink
                    elif diff < 180:  return "FFFFA500", "000000"  # 5-6mo futur → orange
                    else:             return None, None              # >6mo futur → rien
                except ValueError:
                    continue
            return None, None

        def truncate_description(val, max_instances=4):
            """Keep only the first N service instances separated by //"""
            s = str(val or "").strip()
            if not s or s == "nan":
                return ""
            parts = [p.strip() for p in s.split("//")]
            kept = parts[:max_instances]
            result = " // ".join(kept)
            if len(parts) > max_instances:
                result += f" … (+{len(parts)-max_instances} more)"
            return result

        col_labels = {
            "date": "Date", "technicien": "Technicien", "sequence": "Seq",
            "job_id": "Job ID", "cust": "Client", "duo": "DUO", "ot": "OT",
            "debut": "Début", "fin": "Fin", "adresse": "Adresse",
            "travel_min": "Trajet (min)", "job_min": "Job (min)", "buffer_min": "Buffer (min)",
            "techs_needed": "Techs", "unit": "Unité", "serial_number": "Série",
            "difference": "Différence", "description": "Description",
            "insp_wo": "Insp. WO#", "insp_type": "Insp. Type", "insp_date": "Insp. Date",
            "job_minutes": "Durée (min)", "techs_needed": "Techs",
            "postal": "Code postal", "ot_impossible": "OT impossible",
            "address": "Adresse","all_open_work": "All Open Work",
        }

        col_widths = {
            "date": 12, "technicien": 22, "sequence": 5, "job_id": 14,
            "cust": 10, "duo": 6, "ot": 6, "debut": 8, "fin": 8,
            "adresse": 35, "address": 35, "travel_min": 10, "job_min": 10,
            "buffer_min": 10, "techs_needed": 7, "unit": 10,
            "serial_number": 14, "difference": 12, "description": 45,
            "insp_wo": 14, "insp_type": 18, "insp_date": 14,
            "job_minutes": 12, "postal": 12, "ot_impossible": 12,"all_open_work": 40,
        }

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", fgColor="FF2C3E50")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def write_sheet(ws, data_df, is_planning=True, row_offset=0):
            out = data_df.copy()

            # Split last_inspection
            if "last_inspection" in out.columns:
                parsed = out["last_inspection"].apply(parse_last_insp)
                loc = out.columns.get_loc("last_inspection")
                out.insert(loc,     "insp_wo",   parsed.apply(lambda x: x[0]))
                out.insert(loc + 1, "insp_type", parsed.apply(lambda x: x[1]))
                out.insert(loc + 2, "insp_date", parsed.apply(lambda x: x[2]))
                out = out.drop(columns=["last_inspection"])

            # Truncate description
            if "description" in out.columns:
                out["description"] = out["description"].apply(truncate_description)

            cols = list(out.columns)
            insp_date_col_idx = cols.index("insp_date") + 1 if "insp_date" in cols else None

            # Header
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1 + row_offset, column=ci, value=col_labels.get(col, col.replace("_", " ").title()))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[1 + row_offset].height = 30

            # Data rows
            for ri, (_, row) in enumerate(out.iterrows(), 2 + row_offset):
                is_return_home = is_planning and str(row.get("job_id", "")).upper() == "RETURN_HOME"
                is_ot  = is_planning and bool(str(row.get("ot", "")).strip())
                is_duo = is_planning and bool(str(row.get("duo", "")).strip())

                if is_return_home:
                    row_bg, row_fg = "FFF0F0F0", "888888"
                elif is_ot:
                    row_bg, row_fg = "FFF8D7DA", "000000"
                elif is_duo:
                    row_bg, row_fg = "FFD4EDDA", "000000"
                elif ri % 2 == 0:
                    row_bg, row_fg = "FFFAFAFA", "000000"
                else:
                    row_bg, row_fg = "FFFFFFFF", "000000"

                row_fill = PatternFill(fill_type="solid", fgColor=row_bg)
                row_font = Font(color=row_fg, size=9)

                for ci, col in enumerate(cols, 1):
                    val = row[col]
                    if hasattr(val, "item"): val = val.item()
                    # Colonne différence → entier arrondi (pas de décimale, pas de texte)
                    if col == "difference":
                        try:
                            display_val = int(round(float(val))) if val not in (None, "", "nan", "None") and str(val) not in ("nan", "None") else None
                        except (ValueError, TypeError):
                            display_val = None
                    else:
                        display_val = "" if str(val) in ("nan", "None") else val
                    cell = ws.cell(row=ri, column=ci, value=display_val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=(col == "description"))
                    if col == "difference" and display_val is not None:
                        cell.number_format = '#,##0'

                    if ci == insp_date_col_idx and not is_return_home:
                        bg, fg = date_color(val)
                        if bg:
                            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                            cell.font = Font(color=fg, size=9, bold=True)
                        else:
                            cell.fill = row_fill; cell.font = row_font
                    else:
                        cell.fill = row_fill; cell.font = row_font

            # Column widths
            for ci, col in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 14)

            # Freeze header + first 3 cols for planning
            freeze_row = 2 + row_offset
            ws.freeze_panes = f"D{freeze_row}" if is_planning else f"B{freeze_row}"

        wb = Workbook()

        # ── Sheet 1: Planning ────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Planning"
        write_sheet(ws1, df, is_planning=True)

        # ── Sheet 2: Unplanned jobs ──────────────────────────────────
        if unplanned_df is not None and not unplanned_df.empty:
            ws2 = wb.create_sheet("Jobs non planifiés")
            # Add color legend in first row
            ws2.cell(row=1, column=1, value="🗓️ Jobs non planifiés ce mois")
            ws2.cell(row=1, column=1).font = Font(bold=True, size=12, color="C0392B")
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(6, len(unplanned_df.columns)))

            # Écrire directement dans ws2 en décalant d'une ligne (ligne 1 = titre)
            write_sheet(ws2, unplanned_df.copy(), is_planning=False, row_offset=1)

            ws2.row_dimensions[1].height = 22
            ws2.freeze_panes = "B3"

        from io import BytesIO as _BytesIO
        buf = _BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def render_integrity_check_bookable(initial_ids_set: set):
        # Recalcul uniquement si les rows ont changé (via longueur comme proxy rapide)
        cache_key = "integrity_cache"
        planned_len = len(st.session_state.get("planning_month_rows", []))
        remaining_len = len(st.session_state.get("planning_month_remaining_rows", []))
        current_sig = (planned_len, remaining_len)

        if st.session_state.get(cache_key + "_sig") != current_sig:
            planned_jobs = set()
            for r in st.session_state.get("planning_month_rows", []):
                jid = r.get("job_id", "")
                if jid and str(jid).upper() != "RETURN_HOME":
                    planned_jobs.add(normalize_base_job_id(jid))

            remaining_jobs = set()
            for r in st.session_state.get("planning_month_remaining_rows", []):
                jid = r.get("job_id", "")
                if jid:
                    remaining_jobs.add(normalize_base_job_id(jid))

            st.session_state[cache_key + "_planned"] = planned_jobs
            st.session_state[cache_key + "_remaining"] = remaining_jobs
            st.session_state[cache_key + "_sig"] = current_sig
        else:
            planned_jobs = st.session_state.get(cache_key + "_planned", set())
            remaining_jobs = st.session_state.get(cache_key + "_remaining", set())

        covered_jobs = planned_jobs.union(remaining_jobs)
        missing_jobs = sorted(list(initial_ids_set - covered_jobs))

        st.divider()
        st.subheader("🔎 Vérification d'intégrité des jobs (pool bookable)")
        st.write(f"Jobs bookables initiaux : {len(initial_ids_set)}")
        st.write(f"Jobs planifiés : {len(planned_jobs)}")
        st.write(f"Jobs restants : {len(remaining_jobs)}")
        st.write(f"Jobs couverts total : {len(covered_jobs)}")

        if missing_jobs:
            st.error(f"⚠️ Jobs manquants détectés : {len(missing_jobs)}")
            st.dataframe(pd.DataFrame({"Jobs à investiguer": missing_jobs}), use_container_width=True)
        else:
            st.success("✅ Aucun job perdu (bookable). Tous les jobs sont planifiés ou listés comme restants.")


    # ════════════════════════════════════════════════════════════════
    # MODE SELECTOR
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.subheader("🧭 Mode de planification")
    mode = st.radio(
        "Choisir un mode",
        [
            "1 journée / 1 technicien (mode actuel)",
            "Mois complet — techniciens choisis par l'utilisateur",
            "Mois complet — techniciens choisis automatiquement",
        ],
        horizontal=True,
        key="p2_mode",
    )

    tech_names_all = sorted(tech_df["tech_name"].astype(str).tolist())

    # ════════════════════════════════════════════════════════════════
    # MODE A — 1 journée / 1 technicien
    # ════════════════════════════════════════════════════════════════
    if mode == "1 journée / 1 technicien (mode actuel)":
        st.subheader("🧰 Planning 1 journée / 1 technicien")
        chosen_tech = st.selectbox("Choisir le technicien", tech_names_all, index=0, key="p2_chosen_tech")
        home_addr = tech_df.loc[tech_df["tech_name"] == chosen_tech, "home_address"].iloc[0]
        st.caption(f"🏠 Adresse domicile: {home_addr}")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            day_hours = st.number_input("Heures/jour", 4.0, 14.0, 8.0, 0.5, key="p2_day_hours")
        with c2:
            lunch_min = st.number_input("Pause (min)", 0, 120, 30, 5, key="p2_lunch")
        with c3:
            buffer_job = st.number_input("Buffer/job (min)", 0, 60, 10, 5, key="p2_buffer")
        with c4:
            max_jobs = st.number_input("Max jobs/jour", 1, 25, 10, 1, key="p2_max_jobs")

        only_one = st.checkbox("Filtrer: seulement jobs à 1 technicien", value=False, key="p2_only_one")
        service_choice = st.selectbox(
            "Type de jobs",
            ["Inclure full service (tous les jobs)", "Generator inspection seulement", "Exclure generator inspection"],
            index=0, key="p2_service_choice",
        )

        run = st.button("🚀 Générer la journée", type="primary", key="p2_run")

        if run:
            st.session_state["p2_api_calls"] = 0
            st.session_state["p2_cache_hits"] = 0

            available = int(round(day_hours * 60)) - int(lunch_min)
            daily_onsite_cap = int(available)
            OT_ACTIVE_CAP = int(round(14 * 60)) - int(lunch_min)
            if OT_ACTIVE_CAP < available:
                OT_ACTIVE_CAP = available
            MIN_ONSITE_CHUNK_MIN = 180

            remaining = jobs.copy()
            remaining = filter_by_service_type(remaining, service_choice)
            if only_one:
                remaining = remaining[remaining["techs_needed"] <= 1].copy()
            remaining = remaining.sort_values(["job_id"], kind="mergesort")

            used = 0
            seq = 0
            cur_loc = home_addr
            day_rows = []
            carryover = None
            tsec = tech_sector_map.get(chosen_tech, "UNK")

            while True:
                if remaining.empty:
                    break
                if seq >= int(max_jobs):
                    break
                if carryover is not None and int(carryover.get("remaining_job_min", 0)) > 0:
                    break

                best_idx = None
                best_cost = None
                best_tmin = None

                sample = get_job_pool_for_tech(remaining, chosen_tech, int(solo_pool))

                for idx, job in sample.iterrows():
                    jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                    if not sector_compatible(tsec, jsec):
                        continue
                    tlat_ma, tlon_ma = tech_ll_map.get(chosen_tech, (None, None)) if "chosen_tech" in dir() else (None, None)
                    jlat_ma = jobs.at[idx, "job_lat"] if idx in jobs.index and "job_lat" in jobs.columns else None
                    jlon_ma = jobs.at[idx, "job_lon"] if idx in jobs.index and "job_lon" in jobs.columns else None
                    tmin = travel_min_estimate(cur_loc, job["address"], tlat_ma, tlon_ma, jlat_ma, jlon_ma)
                    tback = travel_min_estimate(job["address"], home_addr, jlat_ma, jlon_ma, tlat_ma, tlon_ma)
                    need = int(tmin) + int(job["job_minutes"]) + int(buffer_job) + int(tback)
                    if need <= 0:
                        continue
                    if used + need <= available:
                        if best_cost is None or int(tmin) < best_cost:
                            best_idx = idx
                            best_cost = int(tmin)
                            best_tmin = int(tmin)

                if best_idx is not None:
                    job = jobs.loc[best_idx] if best_idx in jobs.index else remaining.loc[best_idx]
                    seq += 1
                    start_m = used + int(best_tmin)
                    end_m = start_m + int(job["job_minutes"]) + int(buffer_job)
                    day_rows.append({
                        "technicien": chosen_tech, "sequence": seq,
                        "job_id": job["job_id"], "cust": job.get("cust", ""),
                        "duo": "⚠️ DUO" if int(job["techs_needed"]) >= 2 else "",
                        "ot": "", "debut": mm_to_hhmm(int(start_m)), "fin": mm_to_hhmm(int(end_m)),
                        "adresse": job["address"], "travel_min": int(best_tmin),
                        "job_min": int(job["job_minutes"]), "buffer_min": int(buffer_job),
                        "techs_needed": int(job["techs_needed"]), "description": job["description"],
                        "last_inspection": job.get("last_inspection", ""),
                        "difference": job.get("difference", ""),
                        "unit": job.get("unit", ""),
                        "serial_number": job.get("serial_number", ""),
                    })
                    used = int(end_m)
                    cur_loc = job["address"]
                    remaining = remaining[remaining["job_id"] != job["job_id"]].copy()
                    remaining = remaining.sort_values(["job_id"], kind="mergesort")
                    continue

                if seq == 0:
                    best_ot_idx = None
                    best_ot_cost = None
                    best_ot_tmin = None
                    for idx, job in sample.iterrows():
                        jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                        if not sector_compatible(tsec, jsec):
                            continue
                        _tlat_a, _tlon_a = tech_ll_map.get(chosen_tech, (None, None))
                        tmin = travel_min_estimate(cur_loc, job["address"], _tlat_a, _tlon_a, jlat, jlon)
                        tback = travel_min_estimate(job["address"], home_addr, jlat, jlon, _tlat_a, _tlon_a)
                        need = int(tmin) + int(job["job_minutes"]) + int(buffer_job) + int(tback)
                        if need <= OT_ACTIVE_CAP:
                            if best_ot_cost is None or int(tmin) < best_ot_cost:
                                best_ot_idx = idx
                                best_ot_cost = int(tmin)
                                best_ot_tmin = int(tmin)

                    if best_ot_idx is not None:
                        job = jobs.loc[best_ot_idx] if best_ot_idx in jobs.index else remaining.loc[best_ot_idx]
                        seq += 1
                        start_m = used + int(best_ot_tmin)
                        end_m = start_m + int(job["job_minutes"]) + int(buffer_job)
                        day_rows.append({
                            "technicien": chosen_tech, "sequence": seq,
                            "job_id": job["job_id"], "cust": job.get("cust", ""),
                            "duo": "", "ot": "🟥 OT",
                            "debut": mm_to_hhmm(int(start_m)), "fin": mm_to_hhmm(int(end_m)),
                            "adresse": job["address"], "travel_min": int(best_ot_tmin),
                            "job_min": int(job["job_minutes"]), "buffer_min": int(buffer_job),
                            "techs_needed": int(job["techs_needed"]), "description": job["description"],
                        })
                        used = int(end_m)
                        cur_loc = job["address"]
                        remaining = remaining[remaining["job_id"] != job["job_id"]].copy()
                    break

                # Split long jobs
                best_long = None
                for idx, job in sample.iterrows():
                    jm = int(job["job_minutes"])
                    if jm <= int(daily_onsite_cap):
                        continue
                    jlat, jlon, jsec = ensure_job_ll_master(jobs, idx) if idx in jobs.index else (*get_ll_for_address(job.get("address","")), classify_sector(*get_ll_for_address(job.get("address",""))))
                    if not sector_compatible(tsec, jsec):
                        continue
                    tlat_ma, tlon_ma = tech_ll_map.get(chosen_tech, (None, None)) if "chosen_tech" in dir() else (None, None)
                    jlat_ma = jobs.at[idx, "job_lat"] if idx in jobs.index and "job_lat" in jobs.columns else None
                    jlon_ma = jobs.at[idx, "job_lon"] if idx in jobs.index and "job_lon" in jobs.columns else None
                    tmin = travel_min_estimate(cur_loc, job["address"], tlat_ma, tlon_ma, jlat_ma, jlon_ma)
                    tback = travel_min_estimate(job["address"], home_addr, jlat_ma, jlon_ma, tlat_ma, tlon_ma)
                    max_onsite_today = int(available) - int(used) - int(tmin) - int(buffer_job) - int(tback)
                    if max_onsite_today <= 0:
                        continue
                    onsite_today_candidate = choose_onsite_no_crumbs(jm, max_onsite_today, MIN_ONSITE_CHUNK_MIN)
                    if onsite_today_candidate <= 0:
                        continue
                    if best_long is None or int(tmin) < int(best_long["tmin"]):
                        best_long = {"idx": idx, "job": job, "tmin": int(tmin), "tback": int(tback)}

                if best_long is None:
                    break

                job = jobs.loc[best_long["idx"]] if best_long["idx"] in jobs.index else remaining.loc[best_long["idx"]]
                base_job_id = str(job["job_id"])
                total_parts = compute_total_parts(int(job["job_minutes"]), int(daily_onsite_cap))
                carryover = {
                    "base_job_id": base_job_id, "cust": job.get("cust", ""),
                    "address": job["address"], "description": job["description"],
                    "techs_needed": int(job.get("techs_needed", 1)),
                    "total_parts": int(total_parts), "part_idx_next": 1,
                    "remaining_job_min": int(job["job_minutes"]),
                }
                remaining = remaining[remaining["job_id"] != job["job_id"]].copy()

                tmin = best_long["tmin"]
                tback = best_long["tback"]
                max_onsite_today = int(available) - int(used) - int(tmin) - int(buffer_job) - int(tback)
                onsite_today = choose_onsite_no_crumbs(int(carryover["remaining_job_min"]), int(max_onsite_today), MIN_ONSITE_CHUNK_MIN)
                if onsite_today <= 0:
                    break

                seq += 1
                start_m = int(used) + int(tmin)
                end_m = start_m + int(onsite_today) + int(buffer_job)
                day_rows.append({
                    "technicien": chosen_tech, "sequence": seq,
                    "job_id": f"{base_job_id} (PART 1/{total_parts})",
                    "cust": carryover.get("cust", ""), "duo": "", "ot": "",
                    "debut": mm_to_hhmm(int(start_m)), "fin": mm_to_hhmm(int(end_m)),
                    "adresse": job["address"], "travel_min": int(tmin),
                    "job_min": int(onsite_today), "buffer_min": int(buffer_job),
                    "techs_needed": int(job.get("techs_needed", 1)), "description": job["description"],
                })
                used = int(end_m)
                cur_loc = job["address"]
                carryover["remaining_job_min"] = int(carryover["remaining_job_min"]) - int(onsite_today)
                carryover["part_idx_next"] = 2
                break

            st.session_state["planning_day_rows"] = day_rows
            st.session_state["planning_remaining_count"] = len(remaining)

        day_rows_saved = st.session_state.get("planning_day_rows", [])
        if day_rows_saved:
            st.divider()
            st.subheader("📋 Horaire de la journée (persistant)")
            day_df = pd.DataFrame(day_rows_saved)
            day_df = day_df.sort_values(["technicien", "sequence", "debut"], ascending=True).reset_index(drop=True)
            preferred = ["technicien", "sequence", "job_id", "cust", "duo", "ot", "debut", "fin", "adresse",
                         "travel_min", "job_min", "buffer_min", "techs_needed", "unit", "serial_number", "difference", "last_inspection", "description"]
            cols = [c for c in preferred if c in day_df.columns] + [c for c in day_df.columns if c not in preferred]
            day_df = day_df[cols]
            st.dataframe(style_duo(day_df), use_container_width=True)

            available_active = int(round(st.session_state.get("p2_day_hours", 8.0) * 60)) - int(st.session_state.get("p2_lunch", 30))
            total_travel = int(day_df["travel_min"].sum()) if "travel_min" in day_df.columns else 0
            total_job = int(day_df["job_min"].sum()) if "job_min" in day_df.columns else 0
            total_buffer = int(day_df["buffer_min"].sum()) if "buffer_min" in day_df.columns else 0
            total_active = total_travel + total_job + total_buffer

            st.subheader("📊 Résumé")
            st.write(f"**Total travel:** {total_travel} min")
            st.write(f"**Total job:** {total_job} min")
            st.write(f"**Total buffer:** {total_buffer} min")
            st.write(f"**Total actif:** {total_active} / {available_active} min")
            st.subheader("🧾 Coûts (estimation)")
            st.write(f"**Appels API (cette run):** {st.session_state.get('p2_api_calls', 0)}")
            st.write(f"**Hits cache (cette run):** {st.session_state.get('p2_cache_hits', 0)}")
            st.subheader("🧩 Jobs non planifiés")
            st.caption(f"Reste (approx): {st.session_state.get('planning_remaining_count', '—')} job(s)")

    # ════════════════════════════════════════════════════════════════
    # MODE B — Mois complet, techs imposés
    # ════════════════════════════════════════════════════════════════
    elif mode == "Mois complet — techniciens choisis par l'utilisateur":
        st.subheader("🗓️ Mois complet — techniciens choisis par l'utilisateur")
        month_start, month_end = month_selector("p2m_fixed")
        chosen_techs = st.multiselect(
            "Choisir les techniciens",
            options=tech_names_all,
            default=st.session_state.get("p2_month_fixed_techs", []),
            key="p2_month_fixed_techs",
        )

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            day_hours_m = st.number_input("Heures/jour", 4.0, 14.0, 8.0, 0.5, key="p2m_day_hours")
        with c2:
            lunch_min_m = st.number_input("Pause (min)", 0, 120, 30, 5, key="p2m_lunch")
        with c3:
            buffer_job_m = st.number_input("Buffer/job (min)", 0, 60, 10, 5, key="p2m_buffer")
        with c4:
            max_jobs_m = st.number_input("Max jobs/jour/tech", 1, 25, 10, 1, key="p2m_max_jobs")

        allow_duo = st.checkbox("Autoriser booking DUO (techs_needed = 2)", value=True, key="p2m_allow_duo")

        st.sidebar.subheader("🛠️ Repair (post-planification)")
        do_repair = st.sidebar.checkbox("Activer repair pass", value=True, key="p2_repair_on")
        repair_threshold = st.sidebar.number_input("Seuil travel (min) pour réparer", min_value=20, max_value=120, value=50, step=5, key="p2_repair_thr")
        repair_top_n = st.sidebar.slider("Nb techs candidats (lat/lon)", 2, 6, 4, 1, key="p2_repair_topn")
        repair_max_moves = st.sidebar.slider("Max déplacements (moves)", 0, 80, 25, 5, key="p2_repair_moves")
        repair_timeout = st.sidebar.number_input("Timeout repair (secondes)", min_value=10, max_value=600, value=30, step=10, key="p2_repair_timeout")

        run_month = st.button("🚀 Générer le mois (techs imposés)", type="primary", key="p2_run_month_fixed")

        if run_month:
            st.session_state["p2_api_calls"] = 0
            st.session_state["p2_cache_hits"] = 0

            if len(chosen_techs) == 0:
                st.error("Choisis au moins 1 technicien.")
            else:
                if allow_duo and (jobs["techs_needed"] == 2).any() and len(chosen_techs) < 2:
                    st.error("Il y a des jobs DUO (2 techs), mais tu as sélectionné moins de 2 techniciens.")
                    st.stop()

                days = business_days(month_start, month_end)
                progress = st.progress(0)
                progress_text = st.empty()

                result = schedule_month_with_duo(
                    jobs_in=jobs, tech_names=chosen_techs, month_days=days,
                    day_hours=day_hours_m, lunch_min=lunch_min_m, buffer_job=buffer_job_m,
                    max_jobs_per_day=max_jobs_m, allow_duo=allow_duo,
                    progress=progress, progress_text=progress_text,
                )

                if do_repair and result and result.get("rows"):
                    day_available_min = int(round(day_hours_m * 60)) - int(lunch_min_m)
                    repaired_rows, rep_stats = repair_month_plan(
                        planned_rows=result["rows"], tech_names=chosen_techs,
                        day_available_min=day_available_min, buffer_job=int(buffer_job_m),
                        max_jobs_per_day=int(max_jobs_m), threshold_travel_min=int(repair_threshold),
                        candidate_techs_top_n=int(repair_top_n), max_moves=int(repair_max_moves),
                        timeout_seconds=int(repair_timeout),
                    )
                    result["rows"] = repaired_rows
                    timeout_msg = " ⏱️ timeout" if rep_stats.get("timeout") else ""
                    st.sidebar.caption(f"Repair: moves={rep_stats['moves']} | attempts={rep_stats['attempts']} | improved={rep_stats['improved']}{timeout_msg}")

                # ── Déduplication finale: un job = une seule ligne
                # Passe 1: dédupliquer les jobs (pas RETURN_HOME)
                _dedup_seen: set = set()
                _dedup_rows = []
                for _r in result["rows"]:
                    _jid = str(_r.get("job_id", ""))
                    if _jid.upper() == "RETURN_HOME":
                        continue  # on retraitera les RETURN_HOME après
                    _base = normalize_base_job_id(_jid)
                    _is_split = bool(re.search(r"\(PART\s+\d+/\d+\)", _jid))
                    _is_duo = bool(str(_r.get("duo", "")).strip())
                    _key = (_r.get("date",""), _r.get("technicien",""), _base) if (_is_split or _is_duo) else _base
                    if _key not in _dedup_seen:
                        _dedup_seen.add(_key)
                        _dedup_rows.append(_r)
                # Passe 2: ne garder RETURN_HOME que si le tech a ≥1 job ce jour-là
                _active_dt = set(
                    (_r.get("date",""), _r.get("technicien","")) for _r in _dedup_rows
                )
                for _r in result["rows"]:
                    if str(_r.get("job_id","")).upper() == "RETURN_HOME":
                        if (_r.get("date",""), _r.get("technicien","")) in _active_dt:
                            _dedup_rows.append(_r)
                result["rows"] = _dedup_rows

                st.session_state["planning_month_rows"] = result["rows"]
                st.session_state["planning_month_success"] = result["success"]
                st.session_state["planning_month_mode"] = "fixed"
                st.session_state["planning_month_techs_used"] = chosen_techs

                remaining_df = result["remaining"].copy()
                cols_show = ["job_id", "cust", "address", "description", "job_minutes", "techs_needed", "postal", "ot_impossible"]
                cols_show = [c for c in cols_show if c in remaining_df.columns]
                remaining_show = remaining_df[cols_show].copy() if cols_show else remaining_df.copy()
                st.session_state["planning_month_remaining_rows"] = remaining_show.to_dict("records")

                progress.progress(100)
                progress_text.write("Terminé ✅")

                timing = result.get("timing", {})
                if timing:
                    st.info(
                        f"⏱️ **Post-planification** — "
                        f"Concat: {timing.get('concat_s','?')}s | "
                        f"OT-flag: {timing.get('ot_flag_s','?')}s | "
                        f"Safety net: {timing.get('safety_s','?')}s | "
                        f"**Total post: {timing.get('total_post_s','?')}s**"
                    )

        month_rows_saved = st.session_state.get("planning_month_rows", [])
        if month_rows_saved and st.session_state.get("planning_month_mode") == "fixed":
            st.divider()
            techs_used = st.session_state.get("planning_month_techs_used", [])
            if st.session_state.get("planning_month_success"):
                st.success(f"Mois complété ✅ | Techs utilisés: {len(techs_used)}")
            else:
                st.error("Impossible de compléter le mois avec le nombre de techniciens choisi ❌")
                st.warning("Ajoute des techniciens ou ajuste les paramètres (heures/jour, max jobs, buffer).")

            month_df = pd.DataFrame(month_rows_saved)
            sort_cols = [c for c in ["date", "technicien", "sequence", "debut"] if c in month_df.columns]
            month_df = month_df.sort_values(sort_cols, ascending=True).reset_index(drop=True)
            preferred = ["date", "technicien", "sequence", "job_id", "cust", "duo", "ot", "debut", "fin", "adresse",
                         "travel_min", "job_min", "buffer_min", "techs_needed", "unit", "serial_number", "last_inspection", "difference", "description"]
            cols = [c for c in preferred if c in month_df.columns] + [c for c in month_df.columns if c not in preferred]
            month_df = month_df[cols]

            st.subheader("📋 Horaire du mois (tableau complet)")
            st.dataframe(style_duo(month_df), use_container_width=True)
            st.subheader("👷 Vue par technicien")
            for tech in sorted(month_df["technicien"].dropna().unique()):
                st.markdown(f"### {tech}")
                sub = month_df[month_df["technicien"] == tech].sort_values(["date", "sequence", "debut"], ascending=True)
                st.dataframe(style_duo(sub), use_container_width=True)

            st.subheader("🧾 Coûts (estimation)")
            st.write(f"**Appels API (cette run):** {st.session_state.get('p2_api_calls', 0)}")
            st.write(f"**Hits cache (cette run):** {st.session_state.get('p2_cache_hits', 0)}")

            st.subheader("🧩 Jobs non planifiés")
            remaining_rows = st.session_state.get("planning_month_remaining_rows", [])
            if remaining_rows:
                unplanned_df = pd.DataFrame(remaining_rows)
                st.dataframe(style_duo(unplanned_df), use_container_width=True)
                if "ot_impossible" in unplanned_df.columns:
                    ot_imp = int((unplanned_df["ot_impossible"] == True).sum())
                    if ot_imp > 0:
                        st.warning(f"⚠️ Jobs impossibles même en OT (14h cap): {ot_imp}")
                if "techs_needed" in unplanned_df.columns:
                    duo_left = int((unplanned_df["techs_needed"].astype(int) == 2).sum())
                    if duo_left > 0:
                        st.warning(f"⚠️ Jobs DUO restants (techs_needed=2): {duo_left}")

            # ── Excel Export ──
            st.subheader("📥 Exporter en Excel")
            _ts = pd.Timestamp.now().strftime("%Y-%m-%dT%H-%M")
            _remaining_rows_dl = st.session_state.get("planning_month_remaining_rows", [])
            _unplanned_dl = pd.DataFrame(_remaining_rows_dl) if _remaining_rows_dl else pd.DataFrame()
            _excel_bytes = build_export_excel(month_df, _unplanned_dl if not _unplanned_dl.empty else None)
            st.download_button(
                label="⬇️ Télécharger le planning (.xlsx)",
                data=_excel_bytes,
                file_name=f"{_ts}_planning.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_fixed"
            )

            render_integrity_check_bookable(initial_bookable_job_ids)

    # ════════════════════════════════════════════════════════════════
    # MODE C — Mois complet, techs automatiques
    # [MOYEN-4] Binary search sur k au lieu de linéaire
    # ════════════════════════════════════════════════════════════════
    else:
        st.subheader("⚙️ Mois complet — techniciens choisis automatiquement")
        month_start, month_end = month_selector("p2m_auto")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            day_hours_m = st.number_input("Heures/jour", 4.0, 14.0, 8.0, 0.5, key="p2a_day_hours")
        with c2:
            lunch_min_m = st.number_input("Pause (min)", 0, 120, 30, 5, key="p2a_lunch")
        with c3:
            buffer_job_m = st.number_input("Buffer/job (min)", 0, 60, 10, 5, key="p2a_buffer")
        with c4:
            max_jobs_m = st.number_input("Max jobs/jour/tech", 1, 25, 10, 1, key="p2a_max_jobs")

        allow_duo = st.checkbox("Autoriser booking DUO (techs_needed = 2)", value=True, key="p2a_allow_duo")

        st.sidebar.subheader("🛠️ Repair (post-planification) — auto")
        do_repair = st.sidebar.checkbox("Activer repair pass (auto)", value=True, key="p2_repair_on_auto")
        repair_threshold = st.sidebar.number_input("Seuil travel (min) pour réparer (auto)", min_value=20, max_value=120, value=50, step=5, key="p2_repair_thr_auto")
        repair_top_n = st.sidebar.slider("Nb techs candidats (lat/lon) (auto)", 2, 6, 4, 1, key="p2_repair_topn_auto")
        repair_max_moves = st.sidebar.slider("Max déplacements (moves) (auto)", 0, 80, 25, 5, key="p2_repair_moves_auto")
        repair_timeout = st.sidebar.number_input("Timeout repair (secondes)", min_value=10, max_value=600, value=30, step=10, key="p2_repair_timeout_auto")

        run_month = st.button("🚀 Générer le mois (auto)", type="primary", key="p2_run_month_auto")

        if run_month:
            st.session_state["p2_api_calls"] = 0
            st.session_state["p2_cache_hits"] = 0

            days = business_days(month_start, month_end)
            outer_progress = st.progress(0)
            outer_text = st.empty()

            # [MOYEN-4] Binary search : min_techs_duo = 2 si jobs DUO, sinon 1
            has_duo_jobs = allow_duo and (jobs["techs_needed"] == 2).any()
            lo = 2 if has_duo_jobs else 1
            hi = len(tech_names_all)
            best = None

            def _run_k(k):
                chosen = tech_names_all[:k]
                result = schedule_month_with_duo(
                    jobs_in=jobs, tech_names=chosen, month_days=days,
                    day_hours=day_hours_m, lunch_min=lunch_min_m, buffer_job=buffer_job_m,
                    max_jobs_per_day=max_jobs_m, allow_duo=allow_duo,
                )
                if do_repair and result and result.get("rows"):
                    day_available_min = int(round(day_hours_m * 60)) - int(lunch_min_m)
                    repaired_rows, rep_stats = repair_month_plan(
                        planned_rows=result["rows"], tech_names=chosen,
                        day_available_min=day_available_min, buffer_job=int(buffer_job_m),
                        max_jobs_per_day=int(max_jobs_m), threshold_travel_min=int(repair_threshold),
                        candidate_techs_top_n=int(repair_top_n), max_moves=int(repair_max_moves),
                        timeout_seconds=int(repair_timeout),
                    )
                    result["rows"] = repaired_rows
                    timeout_msg = " ⏱️ timeout" if rep_stats.get("timeout") else ""
                    st.sidebar.caption(f"Repair(auto,k={k}): moves={rep_stats['moves']}{timeout_msg}")
                return result, chosen

            # Phase 1 : binary search pour trouver le k minimal suffisant
            outer_text.write(f"Recherche du nombre optimal de techniciens (binary search)…")
            iterations = 0
            while lo < hi:
                mid = (lo + hi) // 2
                outer_text.write(f"Essai avec {mid} technicien(s)… (recherche {lo}–{hi})")
                result, chosen = _run_k(mid)
                best = {"rows": result["rows"], "success": result["success"],
                        "remaining": result["remaining"], "techs_used": chosen}
                outer_progress.progress(int(((iterations + 1) / max(1, math.ceil(math.log2(hi - lo + 2)))) * 100))
                iterations += 1
                if result["success"]:
                    hi = mid
                else:
                    lo = mid + 1

            # Phase 2 : run final avec lo
            if best is None or not best["success"]:
                outer_text.write(f"Run final avec {lo} technicien(s)…")
                result, chosen = _run_k(lo)
                best = {"rows": result["rows"], "success": result["success"],
                        "remaining": result["remaining"], "techs_used": chosen}

            outer_progress.progress(100)

            # Afficher timing du dernier run
            if best and best.get("timing"):
                timing = best["timing"]
                st.info(
                    f"⏱️ **Post-planification** — "
                    f"Concat: {timing.get('concat_s','?')}s | "
                    f"OT-flag: {timing.get('ot_flag_s','?')}s | "
                    f"Safety net: {timing.get('safety_s','?')}s | "
                    f"**Total post: {timing.get('total_post_s','?')}s**"
                )

            # ── Déduplication finale mode auto
            if best and best.get("rows"):
                _dedup_seen2: set = set()
                _dedup_rows2 = []
                for _r in best["rows"]:
                    _jid = str(_r.get("job_id", ""))
                    if _jid.upper() == "RETURN_HOME":
                        continue
                    _base = normalize_base_job_id(_jid)
                    _is_split = bool(re.search(r"\(PART\s+\d+/\d+\)", _jid))
                    _is_duo = bool(str(_r.get("duo", "")).strip())
                    _key = (_r.get("date",""), _r.get("technicien",""), _base) if (_is_split or _is_duo) else _base
                    if _key not in _dedup_seen2:
                        _dedup_seen2.add(_key)
                        _dedup_rows2.append(_r)
                _active_dt2 = set(
                    (_r.get("date",""), _r.get("technicien","")) for _r in _dedup_rows2
                )
                for _r in best["rows"]:
                    if str(_r.get("job_id","")).upper() == "RETURN_HOME":
                        if (_r.get("date",""), _r.get("technicien","")) in _active_dt2:
                            _dedup_rows2.append(_r)
                best["rows"] = _dedup_rows2
            st.session_state["planning_month_rows"] = best["rows"] if best else []
            st.session_state["planning_month_success"] = best["success"] if best else False
            st.session_state["planning_month_mode"] = "auto"
            st.session_state["planning_month_techs_used"] = best["techs_used"] if best else []

            remaining_df = best["remaining"].copy() if best else pd.DataFrame()
            cols_show = ["job_id", "cust", "address", "description", "job_minutes", "techs_needed", "postal", "ot_impossible"]
            cols_show = [c for c in cols_show if c in remaining_df.columns]
            remaining_show = remaining_df[cols_show].copy() if (not remaining_df.empty and cols_show) else remaining_df.copy()
            st.session_state["planning_month_remaining_rows"] = remaining_show.to_dict("records") if not remaining_show.empty else []

            outer_text.write("Terminé ✅")

        month_rows_saved = st.session_state.get("planning_month_rows", [])
        if month_rows_saved and st.session_state.get("planning_month_mode") == "auto":
            st.divider()
            techs_used = st.session_state.get("planning_month_techs_used", [])
            if st.session_state.get("planning_month_success"):
                st.success(f"Mois complété ✅ | Techs utilisés: {len(techs_used)}")
            else:
                st.error("Même en mode auto, le mois n'a pas pu être complété ❌")
                st.caption("Souvent dû à: paramètres trop restrictifs, jobs 3+ techs, ou journées pleines.")
                st.write("**Techniciens utilisés:**", ", ".join(techs_used) if techs_used else "—")

            month_df = pd.DataFrame(month_rows_saved)
            sort_cols = [c for c in ["date", "technicien", "sequence", "debut"] if c in month_df.columns]
            month_df = month_df.sort_values(sort_cols, ascending=True).reset_index(drop=True)
            preferred = ["date", "technicien", "sequence", "job_id", "cust", "duo", "ot", "debut", "fin", "adresse",
                         "travel_min", "job_min", "buffer_min", "techs_needed", "unit", "serial_number", "last_inspection", "difference", "description"]
            cols = [c for c in preferred if c in month_df.columns] + [c for c in month_df.columns if c not in preferred]
            month_df = month_df[cols]

            st.subheader("📋 Horaire du mois (tableau complet)")
            st.dataframe(style_duo(month_df), use_container_width=True)
            st.subheader("👷 Vue par technicien")
            for tech in sorted(month_df["technicien"].dropna().unique()):
                st.markdown(f"### {tech}")
                sub = month_df[month_df["technicien"] == tech].sort_values(["date", "sequence", "debut"], ascending=True)
                st.dataframe(style_duo(sub), use_container_width=True)

            st.subheader("🧾 Coûts (estimation)")
            st.write(f"**Appels API (cette run):** {st.session_state.get('p2_api_calls', 0)}")
            st.write(f"**Hits cache (cette run):** {st.session_state.get('p2_cache_hits', 0)}")

            st.subheader("🧩 Jobs non planifiés")
            remaining_rows = st.session_state.get("planning_month_remaining_rows", [])
            if remaining_rows:
                unplanned_df = pd.DataFrame(remaining_rows)
                st.dataframe(style_duo(unplanned_df), use_container_width=True)
                if "ot_impossible" in unplanned_df.columns:
                    ot_imp = int((unplanned_df["ot_impossible"] == True).sum())
                    if ot_imp > 0:
                        st.warning(f"⚠️ Jobs impossibles même en OT (14h cap): {ot_imp}")
                if "techs_needed" in unplanned_df.columns:
                    duo_left = int((unplanned_df["techs_needed"].astype(int) == 2).sum())
                    if duo_left > 0:
                        st.warning(f"⚠️ Jobs DUO restants (techs_needed=2): {duo_left}")

            # ── Excel Export ──
            st.subheader("📥 Exporter en Excel")
            _ts2 = pd.Timestamp.now().strftime("%Y-%m-%dT%H-%M")
            _remaining_rows_dl2 = st.session_state.get("planning_month_remaining_rows", [])
            _unplanned_dl2 = pd.DataFrame(_remaining_rows_dl2) if _remaining_rows_dl2 else pd.DataFrame()
            _excel_bytes2 = build_export_excel(month_df, _unplanned_dl2 if not _unplanned_dl2.empty else None)
            st.download_button(
                label="⬇️ Télécharger le planning (.xlsx)",
                data=_excel_bytes2,
                file_name=f"{_ts2}_planning.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_auto"
            )

            render_integrity_check_bookable(initial_bookable_job_ids)


# ────────────────────────────────────────────────────────────────
# Router
# ────────────────────────────────────────────────────────────────
if page == "🏠 Route Optimizer":
    render_page_1()

    st.divider()

    with st.expander(
        "🤖 Service Coordinator Assistant",
        expanded=False
    ):
        render_service_assistant()

elif page == "📅 Planning (Page 2)":
    render_page_2()

elif page == "⏱ Feuille de temps":
    show_timesheet()


